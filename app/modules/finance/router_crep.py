# -*- coding: utf-8 -*-
"""
Endpoints de la conciliación con el BCP.

Van aparte del router de finanzas porque son un circuito propio: subir los
reportes del banco, revisar qué se aplicó y descargar el archivo de cobranza.
Todo lo que escribe en la base tiene modo simulación, porque una conciliación
mal aplicada da por pagadas cuotas que nadie pagó.
"""

from __future__ import annotations

import datetime as dt
import io
from decimal import Decimal, InvalidOperation
from typing import List, Optional

from fastapi import (APIRouter, Depends, File, Form, HTTPException, Query,
                     UploadFile, status)
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.exc import ProgrammingError, OperationalError
from sqlalchemy.orm import Session, joinedload

from app.core.util.security import get_current_user
from app.db.database import get_db
from app.modules.finance import conciliacion as con
from app.modules.finance import models as fin
from app.modules.finance.crep import ErrorFormatoBCP

router = APIRouter(prefix="/finance/crep", tags=["Finanzas - BCP"])

TAMANO_MAXIMO = 8 * 1024 * 1024      # 8 MB: el CREP real pesa 0.9 MB
MAXIMO_ARCHIVOS = 40


def _solo_admin(current_user: dict):
    if current_user.get("rol") != "ADMIN":
        raise HTTPException(status_code=403,
                            detail="No puedes acceder a la conciliación de pagos")


def _sin_tablas(e: Exception) -> HTTPException:
    """Las tablas se crean con 12_conciliacion_bcp.sql."""
    return HTTPException(
        status_code=503,
        detail="Falta preparar la base de datos: ejecuta el script "
               "12_conciliacion_bcp.sql para crear las tablas de conciliación.")


async def _leer(archivos: List[UploadFile]) -> List[tuple]:
    if not archivos:
        raise HTTPException(400, "No se recibió ningún archivo")
    if len(archivos) > MAXIMO_ARCHIVOS:
        raise HTTPException(
            400, f"Demasiados archivos de una vez ({len(archivos)}). "
                 f"El máximo es {MAXIMO_ARCHIVOS}")
    salida = []
    for a in archivos:
        datos = await a.read()
        if not datos:
            raise HTTPException(400, f"«{a.filename}» llegó vacío")
        if len(datos) > TAMANO_MAXIMO:
            raise HTTPException(
                400, f"«{a.filename}» pesa {len(datos)/1024/1024:.1f} MB y el "
                     f"máximo es {TAMANO_MAXIMO//1024//1024} MB")
        salida.append((a.filename or "sin_nombre.txt", datos))
    return salida


# ---------------------------------------------------------------------------
# Estado general
# ---------------------------------------------------------------------------

@router.get("/resumen")
def resumen(db: Session = Depends(get_db),
            current_user: dict = Depends(get_current_user)):
    """Lo que hace falta para pintar la pantalla de un vistazo."""
    _solo_admin(current_user)
    try:
        pendientes = (db.query(func.count(fin.Pago.id_pago),
                               func.coalesce(func.sum(fin.Pago.monto), 0),
                               func.coalesce(func.sum(fin.Pago.mora), 0))
                      .filter(fin.Pago.estado.in_(con.PENDIENTES)).one())
        externas = (db.query(func.count(fin.CuotaExterna.id_cuota_externa),
                             func.coalesce(func.sum(fin.CuotaExterna.monto), 0),
                             func.coalesce(func.sum(fin.CuotaExterna.mora), 0))
                    .filter(fin.CuotaExterna.estado == "PENDIENTE").one())
        ultimo = (db.query(fin.LoteCobranza)
                  .filter(fin.LoteCobranza.estado != con.ESTADO_LOTE_INICIAL)
                  .order_by(fin.LoteCobranza.id_lote.desc()).first())
        ultimo_lote = None if not ultimo else {
            "id_lote": ultimo.id_lote,
            "archivo": ultimo.nombre_archivo,
            "fecha_reporte": ultimo.fecha_reporte.isoformat() if ultimo.fecha_reporte else None,
            "fecha_carga": ultimo.fecha_carga.isoformat() if ultimo.fecha_carga else None,
            "aplicados": ultimo.aplicados,
            "sin_coincidencia": ultimo.sin_coincidencia,
        }
        inicial = con.estado_puesta_en_marcha(db)
        sincronizacion = con.obtener_estado_crep_y_cambios(db)
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())

    return {
        "importacion_inicial": inicial,
        "cuotas_pendientes": int(pendientes[0]) + int(externas[0]),
        "de_alumnos": int(pendientes[0]),
        "deuda_historica": int(externas[0]),
        "deuda": float(pendientes[1]) + float(externas[1]),
        "mora": float(pendientes[2]) + float(externas[2]),
        "vencimientos_sin_mora": con.vencimientos_sin_mora(db),
        "cobros_por_revisar": db.query(func.count(fin.MovimientoCobranza.id_movimiento))
            .filter(fin.MovimientoCobranza.estado == "PENDIENTE_REVISION").scalar() or 0,
        "ultimo_lote": ultimo_lote,
        "sincronizacion_crep": sincronizacion,
    }


@router.get("/lotes")
def listar_lotes(limite: int = Query(30, ge=1, le=200),
                 db: Session = Depends(get_db),
                 current_user: dict = Depends(get_current_user)):
    """Historial de reportes procesados, del más reciente al más antiguo."""
    _solo_admin(current_user)
    try:
        # Sin la puesta en marcha: no es un reporte de cobros y sus cifras
        # significan otra cosa, así que en esta lista solo confundiría.
        filas = (db.query(fin.LoteCobranza)
                 .filter(fin.LoteCobranza.estado != con.ESTADO_LOTE_INICIAL)
                 .order_by(fin.LoteCobranza.id_lote.desc()).limit(limite).all())
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())
    return [{
        "id_lote": l.id_lote,
        "archivo": l.nombre_archivo,
        "fecha_reporte": l.fecha_reporte.isoformat() if l.fecha_reporte else None,
        "fecha_carga": l.fecha_carga.isoformat() if l.fecha_carga else None,
        "registros": l.registros_declarados,
        "monto": float(l.monto_declarado or 0),
        "aplicados": l.aplicados,
        "sin_coincidencia": l.sin_coincidencia,
        "extornados": l.extornados,
    } for l in filas]


@router.get("/lotes/{id_lote}/movimientos")
def movimientos_del_lote(id_lote: int,
                         resultado: Optional[str] = None,
                         db: Session = Depends(get_db),
                         current_user: dict = Depends(get_current_user)):
    """Las líneas de un reporte y qué se hizo con cada una."""
    _solo_admin(current_user)
    lote = db.query(fin.LoteCobranza).filter(
        fin.LoteCobranza.id_lote == id_lote).first()
    if not lote:
        raise HTTPException(404, "Ese lote no existe")

    consulta = (db.query(fin.MovimientoCobranza)
                .options(joinedload(fin.MovimientoCobranza.pago)
                         .joinedload(fin.Pago.alumno))
                .filter(fin.MovimientoCobranza.id_lote == id_lote))
    if resultado:
        consulta = consulta.filter(fin.MovimientoCobranza.resultado == resultado.upper())

    return [{
        "id_movimiento": m.id_movimiento,
        "documento": m.documento,
        "alumno": m.pago.alumno_nombre if m.pago else None,
        "fecha_vencimiento": m.fecha_vencimiento.isoformat() if m.fecha_vencimiento else None,
        "fecha_pago": m.fecha_pago.isoformat() if m.fecha_pago else None,
        "monto_pagado": float(m.monto_pagado or 0),
        "mora_pagada": float(m.mora_pagada or 0),
        "monto_total": float(m.monto_total or 0),
        "operacion": m.operacion,
        "medio": m.medio_atencion,
        "resultado": m.resultado,
        "detalle": m.detalle,
    } for m in consulta.all()]


# ---------------------------------------------------------------------------
# Procesar reportes de cobros
# ---------------------------------------------------------------------------

@router.post("/reportes")
async def cargar_reportes(archivos: List[UploadFile] = File(...),
                          simular: bool = Form(True),
                          db: Session = Depends(get_db),
                          current_user: dict = Depends(get_current_user)):
    """Sube uno o varios 'Reporte de cobros' del BCP.

    Se ordenan solos por la fecha que trae cada archivo, no por el orden en
    que se suban: aplicar agosto antes que julio dejaría mora a alumnos que ya
    habían pagado.

    Con `simular` en true no se escribe nada; sirve para ver qué pasaría.
    """
    _solo_admin(current_user)
    lista = await _leer(archivos)
    try:
        return con.procesar_reportes(db, lista,
                                     id_usuario=current_user.get("id"),
                                     simular=simular)
    except ErrorFormatoBCP as e:
        db.rollback()
        raise HTTPException(400, str(e))
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"No se pudo procesar la carga: {e}")


@router.post("/importacion-inicial")
async def importacion_inicial(archivo: UploadFile = File(...),
                              simular: bool = Form(True),
                              sincronizar_pagadas: bool = Form(True),
                              db: Session = Depends(get_db),
                              current_user: dict = Depends(get_current_user)):
    """Toma el CREP que el colegio usa hoy y alinea la base con él.

    Se hace UNA vez, al arrancar el módulo. Da de alta la deuda de quien ya no
    está matriculado, iguala la mora a la del archivo y, si se pide, marca como
    pagadas las cuotas que el archivo ya no trae.
    """
    _solo_admin(current_user)
    (nombre, datos), = await _leer([archivo])
    try:
        return con.importar_crep_inicial(db, datos, nombre, simular=simular,
                                         sincronizar_pagadas=sincronizar_pagadas)
    except ErrorFormatoBCP as e:
        db.rollback()
        raise HTTPException(400, str(e))
    except ValueError as e:
        # Reaplicar el mismo archivo de puesta en marcha.
        db.rollback()
        raise HTTPException(409, str(e))
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"No se pudo importar el archivo: {e}")


@router.post("/ajustar-importe")
def ajustar_importe(tipo: str = Form(...),
                    id_cuota: int = Form(...),
                    monto: str = Form(...),
                    db: Session = Depends(get_db),
                    current_user: dict = Depends(get_current_user)):
    """Deja en una cuota el importe que trae el archivo del BCP.

    Se llama desde la fila del aviso «importe distinto», una a una. OJO: esto
    sí escribe, no hay simulación. Es a propósito: llega aquí quien ya comparó
    los dos importes y decidió cuál es el bueno.
    """
    _solo_admin(current_user)
    try:
        valor = Decimal(monto)
    except (InvalidOperation, ValueError):
        raise HTTPException(400, f"«{monto}» no es un importe válido")
    try:
        return con.ajustar_importe(db, tipo, id_cuota, valor)
    except ValueError as e:
        db.rollback()
        raise HTTPException(400, str(e))
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"No se pudo cambiar el importe: {e}")


# ---------------------------------------------------------------------------
# Cobros que hay que revisar a mano
# ---------------------------------------------------------------------------

@router.get("/pendientes")
def cobros_pendientes(limite: int = Query(200, ge=1, le=1000),
                      db: Session = Depends(get_db),
                      current_user: dict = Depends(get_current_user)):
    """Cobros del banco que el sistema no pudo aplicar y nadie ha atendido."""
    _solo_admin(current_user)
    try:
        return con.pendientes_de_revision(db, limite=limite)
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())


@router.get("/pendientes/{id_movimiento}/candidatos")
def candidatos(id_movimiento: int, db: Session = Depends(get_db),
               current_user: dict = Depends(get_current_user)):
    """Cuotas que podrían corresponder a ese cobro, la más probable primero."""
    _solo_admin(current_user)
    try:
        return con.candidatos_para(db, id_movimiento)
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.post("/pendientes/{id_movimiento}/resolver")
def resolver(id_movimiento: int,
             accion: str = Form(...),
             id_pago: Optional[int] = Form(None),
             id_cuota_externa: Optional[int] = Form(None),
             nota: Optional[str] = Form(None),
             db: Session = Depends(get_db),
             current_user: dict = Depends(get_current_user)):
    """Cierra a mano un cobro.

    Con `accion=aplicar` la cuota indicada queda pagada y desaparece del
    siguiente archivo de cobranza. Con `accion=descartar` no se toca ninguna
    cuota: solo deja de salir en la bandeja de pendientes.
    """
    _solo_admin(current_user)
    try:
        return con.resolver_movimiento(
            db, id_movimiento, accion=accion, id_pago=id_pago,
            id_cuota_externa=id_cuota_externa, nota=nota,
            id_usuario=current_user.get("id"))
    except ValueError as e:
        db.rollback()
        raise HTTPException(400, str(e))
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"No se pudo resolver el cobro: {e}")


# ---------------------------------------------------------------------------
# Mora
# ---------------------------------------------------------------------------

@router.post("/mora")
def cargar_mora(fecha_vencimiento: str = Form(...),
                importe: Optional[str] = Form(None),
                simular: bool = Form(True),
                db: Session = Depends(get_db),
                current_user: dict = Depends(get_current_user)):
    """Carga la mora a las cuotas de esa fecha que siguen sin pagarse.

    Nunca la duplica: si ya la tienen, se dejan como están.
    """
    _solo_admin(current_user)
    try:
        fecha = dt.date.fromisoformat(fecha_vencimiento)
    except ValueError:
        raise HTTPException(400, "La fecha debe venir como AAAA-MM-DD")

    valor = None
    if importe not in (None, ""):
        try:
            valor = Decimal(str(importe))
        except (InvalidOperation, ValueError):
            raise HTTPException(400, f"El importe {importe!r} no es un número")
        if valor < 0:
            raise HTTPException(400, "El importe de la mora no puede ser negativo")

    try:
        return con.aplicar_mora(db, fecha, importe=valor, simular=simular)
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"No se pudo aplicar la mora: {e}")


# ---------------------------------------------------------------------------
# Sincronización y Control de Cambios del CREP
# ---------------------------------------------------------------------------

@router.get("/estado-sincronizacion")
def estado_sincronizacion(db: Session = Depends(get_db),
                          current_user: dict = Depends(get_current_user)):
    """Compara el último CREP oficial contra el estado actual de la BD."""
    _solo_admin(current_user)
    try:
        return con.obtener_estado_crep_y_cambios(db)
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())


@router.post("/incorporar-cambios")
def incorporar_cambios(db: Session = Depends(get_db),
                       current_user: dict = Depends(get_current_user)):
    """Incorpora formalmente las bajas/altas al padrón de cobranza oficial del CREP."""
    _solo_admin(current_user)
    try:
        res = con.incorporar_cambios_al_crep(db, id_usuario=current_user.get("id"))
        return res
    except ErrorFormatoBCP as e:
        raise HTTPException(400, str(e))
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())


# ---------------------------------------------------------------------------
# Generación del archivo de cobranza
# ---------------------------------------------------------------------------

@router.get("/vista-previa")
def vista_previa(db: Session = Depends(get_db),
                 current_user: dict = Depends(get_current_user)):
    """Qué llevaría el archivo si se generara ahora, sin descargarlo."""
    _solo_admin(current_user)
    try:
        _, res = con.generar_archivo_crep(db)
    except ErrorFormatoBCP as e:
        raise HTTPException(400, str(e))
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())
    return res


@router.get("/descargar")
def descargar(db: Session = Depends(get_db),
              current_user: dict = Depends(get_current_user)):
    """El archivo CREP, listo para subir al BCP."""
    _solo_admin(current_user)
    try:
        contenido, _ = con.generar_archivo_crep(db)
    except ErrorFormatoBCP as e:
        raise HTTPException(400, str(e))
    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())

    nombre = f"CREP-{dt.date.today().strftime('%d-%m-%Y')}.txt"
    return StreamingResponse(
        io.BytesIO(contenido),
        media_type="text/plain; charset=iso-8859-1",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"',
                 "Content-Length": str(len(contenido))},
    )


@router.get("/descargar-excel")
def descargar_excel(db: Session = Depends(get_db),
                    current_user: dict = Depends(get_current_user)):
    """La misma cobranza en .xlsx, para revisarla o guardarla de respaldo.

    No lleva las macros del BCP: ya no hacen falta, porque el .txt lo genera
    el propio sistema.
    """
    _solo_admin(current_user)
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(
            503, "Falta la librería openpyxl en el servidor para exportar a Excel")

    cuotas = con.cuotas_para_crep(db)
    if not cuotas:
        raise HTTPException(400, "No hay ninguna cuota pendiente que exportar")

    wb = Workbook()
    h = wb.active
    h.title = "Generar Archivo de Cobranza"
    h.append(["Código del Depositante", "Nombre del Depositante",
              "Información de Retorno", "Fecha de Emisión", "Fecha de Vencimiento",
              "Monto a Pagar", "Mora / Cargo Fijo", "Monto Mínimo",
              "Tipo de Registro", "Nro. Documento de Pago",
              "Nro. Documento de Identidad"])
    for c in cuotas:
        h.append([c.codigo_depositante, c.nombre, c.codigo_depositante,
                  c.fecha_emision.strftime("%d/%m/%Y"),
                  c.fecha_vencimiento.strftime("%d/%m/%Y"),
                  float(c.monto), float(c.mora), float(c.monto_minimo),
                  "No Aplica", "RECIBO", c.documento])
    for col, ancho in zip("ABCDEFGHIJK", (22, 42, 22, 16, 18, 14, 16, 14, 15, 20, 24)):
        h.column_dimensions[col].width = ancho
    h.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"Cobranza-{dt.date.today().strftime('%d-%m-%Y')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ---------------------------------------------------------------------------
# Reporte de deudores
# ---------------------------------------------------------------------------

# Excel no admite estos caracteres en el nombre de una hoja, y corta a 31.
_PROHIBIDOS_HOJA = set(r"[]:*?/\'")


def _nombre_hoja(texto: str, usados: set) -> str:
    """Un nombre de hoja válido y único, lo más parecido posible al original."""
    limpio = "".join(" " if c in _PROHIBIDOS_HOJA else c for c in (texto or "")).strip()
    limpio = " ".join(limpio.split()) or "Hoja"
    base = limpio[:31]
    nombre, n = base, 2
    while nombre.lower() in usados:
        # Al recortar, dos secciones distintas pueden chocar. Se numeran.
        sufijo = f" ({n})"
        nombre = base[:31 - len(sufijo)] + sufijo
        n += 1
    usados.add(nombre.lower())
    return nombre


@router.get("/deudores.xlsx")
def deudores_excel(anio: Optional[str] = Query(None, description="Año escolar; por defecto el activo"),
                   hasta: Optional[dt.date] = Query(None, description="Fecha de corte; por defecto hoy"),
                   db: Session = Depends(get_db),
                   current_user: dict = Depends(get_current_user)):
    """Genera el reporte oficial de deudores en Excel con el diseño y estructura exacta:
    
      · Deudores           — Listado general con meses, módulos, deuda exigible, mora, total y deuda acumulada
      · Detalle            — Desglose cuota por cuota para conciliación exacta
      · Resumen por periodo — Tablas dinámicas (Deuda por periodo, Deuda por aula, Periodo por aula) con fórmulas SUMIF/SUMIFS
      · Hojas por sección   — Una hoja formateada por cada aula (Primaria, Secundaria, Sin Aula)
    """
    _solo_admin(current_user)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            503, "Falta la librería openpyxl en el servidor para exportar a Excel")

    hoy = dt.date.today()
    limite = hasta or hoy

    try:
        from app.modules.academic.models import AnioEscolar, Seccion, Grado, Nivel
        from app.modules.enrollment.models import Matricula
        from app.modules.users.relacion_familiar.models import RelacionFamiliar

        # 1. Determinar año escolar
        if not anio:
            activo = (db.query(AnioEscolar).filter(AnioEscolar.activo.is_(True))
                      .order_by(AnioEscolar.id_anio_escolar.desc()).first())
            anio = activo.id_anio_escolar if activo else "2026"

        # 2. Secciones del padrón
        filas_mat = (db.query(Matricula.id_alumno, Seccion.nombre.label("seccion"),
                              Grado.nombre.label("grado"), Grado.orden.label("orden_grado"),
                              Nivel.nombre.label("nivel"))
                     .join(Seccion, Seccion.id_seccion == Matricula.id_seccion)
                     .join(Grado, Grado.id_grado == Seccion.id_grado)
                     .join(Nivel, Nivel.id_nivel == Grado.id_nivel)
                     .filter(Matricula.id_anio_escolar == anio).all())

        def _normalizar_aula(grado_str: str, seccion_str: str, nivel_str: str) -> tuple[str, str]:
            if not grado_str or grado_str == "—" or not seccion_str:
                return "SIN AULA (no esta en el padron)", "—"
            import re
            m_grado = re.search(r'(\d+)', grado_str)
            num_grado = m_grado.group(1) if m_grado else grado_str
            nivel_clean = "Primaria" if "prim" in (nivel_str or "").lower() else "Secundaria" if "sec" in (nivel_str or "").lower() else (nivel_str or "").title()
            grado_texto = f"{num_grado}° {nivel_clean}"
            aula_nombre = f"{num_grado}° {seccion_str.strip()} {nivel_clean}"
            return aula_nombre, grado_texto

        ubicacion: Dict[int, dict] = {}
        for f in filas_mat:
            aula_nom, grado_txt = _normalizar_aula(f.grado, f.seccion, f.nivel)
            ubicacion[f.id_alumno] = {
                "nivel": f.nivel,
                "grado": grado_txt,
                "seccion": f.seccion,
                "aula": aula_nom,
                "orden_grado": f.orden_grado or 0
            }

        # 3. Apoderados y teléfonos
        relaciones = (
            db.query(RelacionFamiliar)
            .options(joinedload(RelacionFamiliar.familiar))
            .all()
        )
        info_apoderados: Dict[int, dict] = {}
        for rel in relaciones:
            f = rel.familiar
            if not f:
                continue
            id_al = rel.id_alumno
            tel = (f.telefono or "").strip()
            nom = (str(f.apellidos or "") + " " + str(f.nombres or "")).strip()
            if not nom:
                nom = (str(f.nombres or "") + " " + str(f.apellidos or "")).strip()
            if id_al in info_apoderados:
                if not info_apoderados[id_al]["telefono"] and tel:
                    info_apoderados[id_al]["telefono"] = tel
                    info_apoderados[id_al]["apoderado"] = nom
                elif tel and tel not in info_apoderados[id_al]["telefono"]:
                    info_apoderados[id_al]["telefono"] += f" / {tel}"
            else:
                info_apoderados[id_al] = {
                    "telefono": tel or "—",
                    "apoderado": nom or "—"
                }

        # 4. Deuda total acumulada en el año
        todos_los_pagos = (db.query(fin.Pago)
                           .filter(fin.Pago.estado.in_(["PENDIENTE", "VENCIDO"]))
                           .all())
        deuda_total_por_alumno: Dict[int, Decimal] = {}
        for p in todos_los_pagos:
            if not p.id_alumno:
                continue
            m = Decimal(str(p.monto or 0)) + Decimal(str(p.mora or 0))
            deuda_total_por_alumno[p.id_alumno] = deuda_total_por_alumno.get(p.id_alumno, Decimal("0")) + m

        # 5. Cuotas exigibles a la fecha de corte
        q_pagos = (db.query(fin.Pago).options(joinedload(fin.Pago.alumno))
                   .filter(fin.Pago.estado.in_(["PENDIENTE", "VENCIDO"]))
                   .filter(fin.Pago.fecha_vencimiento <= limite)
                   .order_by(fin.Pago.fecha_vencimiento.asc(), fin.Pago.id_pago.asc())
                   .all())

        def _categorizar_concepto(concepto_raw: str) -> tuple[str, str]:
            c = (concepto_raw or "").strip()
            c_lower = c.lower()
            meses_map = [
                ("enero", "enero"), ("febrero", "febrero"), ("marzo", "marzo"),
                ("abril", "abril"), ("mayo", "mayo"), ("junio", "junio"),
                ("julio", "julio"), ("agosto", "agosto"), ("septiembre", "septiembre"),
                ("setiembre", "septiembre"), ("octubre", "octubre"), ("noviembre", "noviembre"),
                ("diciembre", "diciembre")
            ]
            for mes_key, mes_norm in meses_map:
                if mes_key in c_lower:
                    import re
                    m_year = re.search(r'\b(20\d\d)\b', c)
                    year_str = f" {m_year.group(1)}" if m_year else f" {anio}"
                    return f"{mes_norm}{year_str}", "Mes"
            if "módulo" in c_lower or "modulo" in c_lower or "material" in c_lower:
                return c, "Módulo"
            if "matrícula" in c_lower or "matricula" in c_lower:
                return c, "Matrícula"
            return c, "Otro"

        SIN_AULA = {"nivel": "—", "grado": "—", "seccion": "SIN SECCIÓN", "aula": "SIN AULA (no esta en el padron)", "orden_grado": 999}
        deudores_dict: Dict[int, dict] = {}
        detalle_cuotas: List[dict] = []

        for p in q_pagos:
            if not p.alumno:
                continue
            monto = Decimal(str(p.monto or 0))
            mora = Decimal(str(p.mora or 0))
            total = monto + mora
            if total <= Decimal("0"):
                continue

            id_al = p.alumno.id_alumno
            donde = ubicacion.get(id_al, SIN_AULA)
            fam = info_apoderados.get(id_al, {"telefono": "—", "apoderado": "—"})

            per_nombre, per_tipo = _categorizar_concepto(p.concepto)
            venc_str = p.fecha_vencimiento.strftime("%d/%m/%Y") if p.fecha_vencimiento else "—"

            cuota_info = {
                "aula": donde["aula"],
                "grado": donde["grado"],
                "dni": p.alumno.dni or "—",
                "alumno": f"{p.alumno.apellidos or ''} {p.alumno.nombres or ''}".strip(),
                "celular": fam["telefono"],
                "apoderado": fam["apoderado"],
                "periodo": per_nombre,
                "tipo": per_tipo,
                "vencimiento": venc_str,
                "vencimiento_date": p.fecha_vencimiento,
                "monto": float(monto),
                "mora": float(mora),
                "total": float(total),
                "id_pago": p.id_pago,
                "en_json": "SI"
            }
            detalle_cuotas.append(cuota_info)

            d = deudores_dict.setdefault(id_al, {
                "id_alumno": id_al,
                "dni": p.alumno.dni or "—",
                "alumno": f"{p.alumno.apellidos or ''} {p.alumno.nombres or ''}".strip(),
                "apoderado": fam["apoderado"],
                "telefono": fam["telefono"],
                "aula": donde["aula"],
                "grado": donde["grado"],
                "nivel": donde["nivel"],
                "orden_grado": donde["orden_grado"],
                "seccion": donde["seccion"],
                "cuotas": [],
                "meses": [],
                "modulos": []
            })
            d["cuotas"].append(cuota_info)
            if per_tipo == "Mes":
                if per_nombre not in d["meses"]:
                    d["meses"].append(per_nombre)
            elif per_tipo in ("Módulo", "Otro"):
                if per_nombre not in d["modulos"]:
                    d["modulos"].append(per_nombre)

        if not deudores_dict:
            raise HTTPException(400, "No hay ninguna deuda pendiente: no hay nada que exportar")

        # Consolidar deudores
        deudores_lista: List[dict] = []
        for id_al, d in deudores_dict.items():
            if not d["cuotas"]:
                continue
            d["cuotas"].sort(key=lambda c: (c["vencimiento_date"] or dt.date.max, c["periodo"]))
            d["num_cuotas"] = len(d["cuotas"])
            d["deuda"] = sum(c["monto"] for c in d["cuotas"])
            d["mora"] = sum(c["mora"] for c in d["cuotas"])
            d["total"] = d["deuda"] + d["mora"]
            d["vencimiento_antiguo"] = d["cuotas"][0]["vencimiento"] if d["cuotas"] else "—"
            d["deuda_total_alumno"] = float(deuda_total_por_alumno.get(id_al, Decimal(str(d["total"]))))
            d["meses_str"] = ", ".join(d["meses"]) if d["meses"] else "—"
            d["modulos_str"] = ", ".join(d["modulos"]) if d["modulos"] else "—"
            deudores_lista.append(d)

        def _aula_sort_key(aula_name: str) -> tuple:
            a_lower = aula_name.lower()
            if "sin aula" in a_lower:
                return (3, 99, 99, aula_name)
            is_prim = "prim" in a_lower
            nivel_order = 1 if is_prim else 2
            import re
            m = re.search(r'(\d+)', aula_name)
            grado_num = int(m.group(1)) if m else 99
            secc_order = 1 if ("amarillo" in a_lower or " a " in a_lower or " a (" in a_lower or a_lower.endswith(" a")) else (2 if ("azul" in a_lower or " b " in a_lower or " b (" in a_lower or a_lower.endswith(" b")) else 3)
            return (nivel_order, grado_num, secc_order, aula_name)

        deudores_lista.sort(key=lambda d: (_aula_sort_key(d["aula"]), d["alumno"]))

        # Agrupar por aula en el orden correcto
        aulas_dict: Dict[str, dict] = {}
        for d in deudores_lista:
            a = aulas_dict.setdefault(d["aula"], {
                "aula": d["aula"],
                "grado": d["grado"],
                "alumnos": []
            })
            a["alumnos"].append(d)

        # Ordenar detalle
        detalle_cuotas.sort(key=lambda c: (_aula_sort_key(c["aula"]), c["alumno"], c["vencimiento_date"] or dt.date.max))

        # Periodos únicos
        periodos_dict: Dict[str, str] = {}
        for c in detalle_cuotas:
            if c["periodo"] not in periodos_dict:
                periodos_dict[c["periodo"]] = c["tipo"]
        periodos_texto = ", ".join(periodos_dict.keys())
        fecha_hora_str = hoy.strftime("%d/%m/%Y")

        # --- ESTILOS VISUALES IDÉNTICOS AL ARCHIVO DE REFERENCIA ---
        FONT_FAMILY = "Calibri"
        TITLE_FONT_FAMILY = "Georgia"

        COLOR_GRANATE = "7F2A19"
        COLOR_CREAM = "E9E1C9"
        COLOR_LIGHT_ZEBRA = "FBF7F0"
        COLOR_TEXT_MAIN = "2E2320"
        COLOR_TEXT_MUTED = "6B5B52"
        COLOR_BORDER = "D3C5B4"
        COLOR_WHITE = "FFFFFF"

        FILL_HEADER = PatternFill(start_color=COLOR_GRANATE, end_color=COLOR_GRANATE, fill_type="solid")
        FILL_BANNER = PatternFill(start_color=COLOR_CREAM, end_color=COLOR_CREAM, fill_type="solid")
        FILL_ZEBRA = PatternFill(start_color=COLOR_LIGHT_ZEBRA, end_color=COLOR_LIGHT_ZEBRA, fill_type="solid")

        FONT_TITLE = Font(name=TITLE_FONT_FAMILY, size=16, bold=True, color=COLOR_GRANATE)
        FONT_SUBTITLE = Font(name=FONT_FAMILY, size=10, bold=False, color=COLOR_TEXT_MUTED)
        FONT_SECTION_HEADER = Font(name=TITLE_FONT_FAMILY, size=12, bold=True, color=COLOR_GRANATE)
        FONT_HEADER = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
        FONT_DATA = Font(name=FONT_FAMILY, size=10, bold=False, color=COLOR_TEXT_MAIN)
        FONT_TOTAL = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_GRANATE)
        FONT_NOTE = Font(name=FONT_FAMILY, size=9, bold=False, color=COLOR_TEXT_MUTED)

        ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

        BORDER_THIN_SIDE = Side(style="thin", color=COLOR_BORDER)
        BORDER_MEDIUM_SIDE = Side(style="medium", color=COLOR_BORDER)

        BORDER_CELL = Border(top=BORDER_THIN_SIDE, bottom=BORDER_THIN_SIDE, left=BORDER_THIN_SIDE, right=BORDER_THIN_SIDE)
        BORDER_TOTAL = Border(top=BORDER_MEDIUM_SIDE, bottom=BORDER_MEDIUM_SIDE, left=BORDER_THIN_SIDE, right=BORDER_THIN_SIDE)

        FORMAT_CURRENCY = '"S/"\\ #,##0.00'

        wb = Workbook()

        # =========================================================================
        # 1. HOJA DEUDORES
        # =========================================================================
        ws_deudores = wb.active
        ws_deudores.title = "Deudores"
        ws_deudores.views.sheetView[0].showGridLines = True

        ws_deudores.merge_cells("A1:N1")
        ws_deudores["A1"] = "LISTA DE ALUMNOS DEUDORES"
        ws_deudores["A1"].font = FONT_TITLE
        ws_deudores["A1"].fill = FILL_BANNER
        ws_deudores["A1"].alignment = ALIGN_CENTER

        ws_deudores.merge_cells("A2:N2")
        ws_deudores["A2"] = f"Periodos solicitados: {periodos_texto}"
        ws_deudores["A2"].font = FONT_SUBTITLE
        ws_deudores["A2"].alignment = ALIGN_LEFT

        ws_deudores.merge_cells("A3:N3")
        ws_deudores["A3"] = f"Generado el {fecha_hora_str}."
        ws_deudores["A3"].font = FONT_SUBTITLE
        ws_deudores["A3"].alignment = ALIGN_LEFT

        headers_deudores = [
            "Aula", "Grado", "Nombre completo del alumno", "DNI estudiante", "Celular / teléfono",
            "Nombre del apoderado", "Meses que debe", "Módulos que debe", "N° de cuotas",
            "Deuda (S/)", "Mora (S/)", "Total a pagar (S/)", "Vencimiento más antiguo", "Deuda total del alumno (S/)"
        ]
        ws_deudores.append([])
        ws_deudores.append(headers_deudores)
        for c in range(1, 15):
            cell = ws_deudores.cell(5, c)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_CELL

        r_idx = 6
        for d in deudores_lista:
            ws_deudores.append([
                d["aula"],
                d["grado"],
                d["alumno"],
                d["dni"],
                d["telefono"],
                d["apoderado"],
                d["meses_str"],
                d["modulos_str"],
                d["num_cuotas"],
                d["deuda"],
                d["mora"],
                f"=J{r_idx}+K{r_idx}",
                d["vencimiento_antiguo"],
                d["deuda_total_alumno"]
            ])
            fill_row = FILL_ZEBRA if r_idx % 2 == 0 else PatternFill(fill_type=None)
            for c in range(1, 15):
                cell = ws_deudores.cell(r_idx, c)
                cell.font = FONT_DATA
                if fill_row.fill_type:
                    cell.fill = fill_row
                cell.border = BORDER_CELL
                if c in (1, 3, 6, 7, 8):
                    cell.alignment = ALIGN_LEFT
                elif c in (2, 4, 5, 9, 13):
                    cell.alignment = ALIGN_CENTER
                elif c in (10, 11, 12, 14):
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = FORMAT_CURRENCY
            r_idx += 1

        total_row_deudores = [
            f"TOTAL ({len(deudores_lista)} alumnos)",
            None, None, None, None, None, None, None,
            f"=SUM(I6:I{r_idx-1})",
            f"=SUM(J6:J{r_idx-1})",
            f"=SUM(K6:K{r_idx-1})",
            f"=SUM(L6:L{r_idx-1})",
            None,
            f"=SUM(N6:N{r_idx-1})"
        ]
        ws_deudores.append(total_row_deudores)
        for c in range(1, 15):
            cell = ws_deudores.cell(r_idx, c)
            cell.font = FONT_TOTAL
            cell.fill = FILL_BANNER
            cell.border = BORDER_TOTAL
            cell.alignment = ALIGN_RIGHT
            if c in (10, 11, 12, 14):
                cell.number_format = FORMAT_CURRENCY

        widths_deudores = {'A': 24.0, 'B': 15.0, 'C': 38.0, 'D': 14.0, 'E': 22.0, 'F': 34.0, 'G': 30.0, 'H': 20.0, 'I': 11.0, 'J': 13.0, 'K': 11.0, 'L': 15.0, 'M': 18.0, 'N': 18.0}
        for col_l, w in widths_deudores.items():
            ws_deudores.column_dimensions[col_l].width = w
        ws_deudores.freeze_panes = "C6"

        # =========================================================================
        # 2. HOJA DETALLE
        # =========================================================================
        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.views.sheetView[0].showGridLines = True

        ws_detalle.merge_cells("A1:M1")
        ws_detalle["A1"] = "DETALLE DE CUOTAS IMPAGAS"
        ws_detalle["A1"].font = FONT_TITLE
        ws_detalle["A1"].fill = FILL_BANNER
        ws_detalle["A1"].alignment = ALIGN_CENTER

        ws_detalle.merge_cells("A2:M2")
        ws_detalle["A2"] = f"Periodos solicitados: {periodos_texto}"
        ws_detalle["A2"].font = FONT_SUBTITLE
        ws_detalle["A2"].alignment = ALIGN_LEFT

        ws_detalle.append([])
        headers_detalle = [
            "Aula", "Grado", "DNI estudiante", "Nombre completo del alumno", "Celular / teléfono",
            "Nombre del apoderado", "Periodo", "Tipo", "Fecha de vencimiento", "Monto (S/)", "Mora (S/)",
            "Fila en hoja de cobranza", "Periodo en periodos.json"
        ]
        ws_detalle.append(headers_detalle)
        for c in range(1, 14):
            cell = ws_detalle.cell(4, c)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_CELL

        det_r = 5
        for c_item in detalle_cuotas:
            ws_detalle.append([
                c_item["aula"],
                c_item["grado"],
                c_item["dni"],
                c_item["alumno"],
                c_item["celular"],
                c_item["apoderado"],
                c_item["periodo"],
                c_item["tipo"],
                c_item["vencimiento"],
                c_item["monto"],
                c_item["mora"],
                c_item["id_pago"],
                c_item["en_json"]
            ])
            fill_row = FILL_ZEBRA if det_r % 2 == 0 else PatternFill(fill_type=None)
            for c in range(1, 14):
                cell = ws_detalle.cell(det_r, c)
                cell.font = FONT_DATA
                if fill_row.fill_type:
                    cell.fill = fill_row
                cell.border = BORDER_CELL
                if c in (1, 4, 6, 7):
                    cell.alignment = ALIGN_LEFT
                elif c in (2, 3, 5, 8, 9, 12, 13):
                    cell.alignment = ALIGN_CENTER
                elif c in (10, 11):
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = FORMAT_CURRENCY
            det_r += 1

        widths_detalle = {'A': 24.0, 'B': 15.0, 'C': 14.0, 'D': 38.0, 'E': 22.0, 'F': 34.0, 'G': 20.0, 'H': 10.0, 'I': 18.0, 'J': 13.0, 'K': 11.0, 'L': 20.0, 'M': 18.0}
        for col_l, w in widths_detalle.items():
            ws_detalle.column_dimensions[col_l].width = w
        ws_detalle.freeze_panes = "C5"
        max_detalle_row = det_r - 1

        # =========================================================================
        # 3. HOJA RESUMEN POR PERIODO
        # =========================================================================
        ws_resumen = wb.create_sheet("Resumen por periodo")
        ws_resumen.views.sheetView[0].showGridLines = True

        ws_resumen.merge_cells("A1:I1")
        ws_resumen["A1"] = "RESUMEN DE DEUDA"
        ws_resumen["A1"].font = FONT_TITLE
        ws_resumen["A1"].fill = FILL_BANNER
        ws_resumen["A1"].alignment = ALIGN_CENTER

        ws_resumen.merge_cells("A2:I2")
        ws_resumen["A2"] = f"Periodos solicitados: {periodos_texto}"
        ws_resumen["A2"].font = FONT_SUBTITLE
        ws_resumen["A2"].alignment = ALIGN_LEFT

        # Tabla 1: Deuda por periodo
        ws_resumen.append([])
        ws_resumen.merge_cells("A4:F4")
        ws_resumen["A4"] = "1.  Deuda por periodo"
        ws_resumen["A4"].font = FONT_SECTION_HEADER
        ws_resumen["A4"].fill = FILL_BANNER
        ws_resumen["A4"].alignment = ALIGN_CENTER

        headers_t1 = ["Periodo", "Tipo", "N° de cuotas impagas", "Deuda (S/)", "Mora (S/)", "Total (S/)"]
        ws_resumen.append(headers_t1)
        for c in range(1, 7):
            cell = ws_resumen.cell(5, c)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_CELL

        r_t1 = 6
        for per, tipo in periodos_dict.items():
            ws_resumen.append([
                per,
                tipo,
                f"=COUNTIF(Detalle!$G$5:$G${max_detalle_row},$A{r_t1})",
                f"=SUMIF(Detalle!$G$5:$G${max_detalle_row},$A{r_t1},Detalle!$J$5:$J${max_detalle_row})",
                f"=SUMIF(Detalle!$G$5:$G${max_detalle_row},$A{r_t1},Detalle!$K$5:$K${max_detalle_row})",
                f"=D{r_t1}+E{r_t1}"
            ])
            fill_row = FILL_ZEBRA if r_t1 % 2 == 1 else PatternFill(fill_type=None)
            for c in range(1, 7):
                cell = ws_resumen.cell(r_t1, c)
                cell.font = FONT_DATA
                if fill_row.fill_type:
                    cell.fill = fill_row
                cell.border = BORDER_CELL
                if c == 1:
                    cell.alignment = ALIGN_LEFT
                elif c in (2, 3):
                    cell.alignment = ALIGN_CENTER
                elif c in (4, 5, 6):
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = FORMAT_CURRENCY
            r_t1 += 1

        ws_resumen.append([
            "TOTAL", None,
            f"=SUM(C6:C{r_t1-1})",
            f"=SUM(D6:D{r_t1-1})",
            f"=SUM(E6:E{r_t1-1})",
            f"=SUM(F6:F{r_t1-1})"
        ])
        for c in range(1, 7):
            cell = ws_resumen.cell(r_t1, c)
            cell.font = FONT_TOTAL
            cell.fill = FILL_BANNER
            cell.border = BORDER_TOTAL
            cell.alignment = ALIGN_RIGHT
            if c in (4, 5, 6):
                cell.number_format = FORMAT_CURRENCY

        # Tabla 2: Deuda por aula
        r_t2_start = r_t1 + 2
        ws_resumen.cell(r_t2_start - 1, 1, "")

        ws_resumen.merge_cells(f"A{r_t2_start}:G{r_t2_start}")
        ws_resumen[f"A{r_t2_start}"] = "2.  Deuda por aula"
        ws_resumen[f"A{r_t2_start}"].font = FONT_SECTION_HEADER
        ws_resumen[f"A{r_t2_start}"].fill = FILL_BANNER
        ws_resumen[f"A{r_t2_start}"].alignment = ALIGN_CENTER

        headers_t2 = ["Aula", "Grado", "Alumnos deudores", "Cuotas impagas", "Deuda (S/)", "Mora (S/)", "Total (S/)"]
        ws_resumen.append(headers_t2)
        for c in range(1, 8):
            cell = ws_resumen.cell(r_t2_start + 1, c)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_CELL

        r_t2 = r_t2_start + 2
        for aula_nom, a_info in aulas_dict.items():
            ws_resumen.append([
                aula_nom,
                a_info["grado"],
                len(a_info["alumnos"]),
                f"=COUNTIF(Detalle!$A$5:$A${max_detalle_row},$A{r_t2})",
                f"=SUMIF(Detalle!$A$5:$A${max_detalle_row},$A{r_t2},Detalle!$J$5:$J${max_detalle_row})",
                f"=SUMIF(Detalle!$A$5:$A${max_detalle_row},$A{r_t2},Detalle!$K$5:$K${max_detalle_row})",
                f"=E{r_t2}+F{r_t2}"
            ])
            fill_row = FILL_ZEBRA if r_t2 % 2 == 1 else PatternFill(fill_type=None)
            for c in range(1, 8):
                cell = ws_resumen.cell(r_t2, c)
                cell.font = FONT_DATA
                if fill_row.fill_type:
                    cell.fill = fill_row
                cell.border = BORDER_CELL
                if c == 1:
                    cell.alignment = ALIGN_LEFT
                elif c in (2, 3, 4):
                    cell.alignment = ALIGN_CENTER
                elif c in (5, 6, 7):
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = FORMAT_CURRENCY
            r_t2 += 1

        ws_resumen.append([
            "TOTAL", None,
            f"=SUM(C{r_t2_start+2}:C{r_t2-1})",
            f"=SUM(D{r_t2_start+2}:D{r_t2-1})",
            f"=SUM(E{r_t2_start+2}:E{r_t2-1})",
            f"=SUM(F{r_t2_start+2}:F{r_t2-1})",
            f"=SUM(G{r_t2_start+2}:G{r_t2-1})"
        ])
        for c in range(1, 8):
            cell = ws_resumen.cell(r_t2, c)
            cell.font = FONT_TOTAL
            cell.fill = FILL_BANNER
            cell.border = BORDER_TOTAL
            cell.alignment = ALIGN_RIGHT
            if c in (5, 6, 7):
                cell.number_format = FORMAT_CURRENCY

        # Tabla 3: Periodo por aula
        r_t3_start = r_t2 + 2
        ws_resumen.cell(r_t3_start - 1, 1, "")

        num_periodos = len(periodos_dict)
        end_col_letter = get_column_letter(num_periodos + 2)
        ws_resumen.merge_cells(f"A{r_t3_start}:{end_col_letter}{r_t3_start}")
        ws_resumen[f"A{r_t3_start}"] = "3.  Periodo por aula   (deuda + mora, en soles)"
        ws_resumen[f"A{r_t3_start}"].font = FONT_SECTION_HEADER
        ws_resumen[f"A{r_t3_start}"].fill = FILL_BANNER
        ws_resumen[f"A{r_t3_start}"].alignment = ALIGN_CENTER

        header_row_t3 = r_t3_start + 1
        headers_t3 = ["Aula", *list(periodos_dict.keys()), "Total del aula"]
        ws_resumen.append(headers_t3)
        for c in range(1, len(headers_t3) + 1):
            cell = ws_resumen.cell(header_row_t3, c)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_CELL

        r_t3 = header_row_t3 + 1
        for aula_nom, a_info in aulas_dict.items():
            row_vals = [aula_nom]
            for p_idx, per in enumerate(periodos_dict.keys(), start=2):
                col_l = get_column_letter(p_idx)
                row_vals.append(
                    f"=SUMIFS(Detalle!$J$5:$J${max_detalle_row},Detalle!$A$5:$A${max_detalle_row},$A{r_t3},Detalle!$G$5:$G${max_detalle_row},{col_l}${header_row_t3})"
                    f"+SUMIFS(Detalle!$K$5:$K${max_detalle_row},Detalle!$A$5:$A${max_detalle_row},$A{r_t3},Detalle!$G$5:$G${max_detalle_row},{col_l}${header_row_t3})"
                )
            first_p_col = "B"
            last_p_col = get_column_letter(num_periodos + 1)
            row_vals.append(f"=SUM({first_p_col}{r_t3}:{last_p_col}{r_t3})")
            ws_resumen.append(row_vals)

            fill_row = FILL_ZEBRA if r_t3 % 2 == 1 else PatternFill(fill_type=None)
            for c in range(1, len(headers_t3) + 1):
                cell = ws_resumen.cell(r_t3, c)
                cell.font = FONT_DATA
                if fill_row.fill_type:
                    cell.fill = fill_row
                cell.border = BORDER_CELL
                if c == 1:
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = FORMAT_CURRENCY
            r_t3 += 1

        total_t3 = ["TOTAL"]
        for p_idx in range(2, num_periodos + 3):
            col_l = get_column_letter(p_idx)
            total_t3.append(f"=SUM({col_l}{header_row_t3+1}:{col_l}{r_t3-1})")
        ws_resumen.append(total_t3)
        for c in range(1, len(headers_t3) + 1):
            cell = ws_resumen.cell(r_t3, c)
            cell.font = FONT_TOTAL
            cell.fill = FILL_BANNER
            cell.border = BORDER_TOTAL
            cell.alignment = ALIGN_RIGHT
            if c > 1:
                cell.number_format = FORMAT_CURRENCY

        ws_resumen.append([])
        ws_resumen.append(["Nota: los periodos marcados 'NO' en la hoja Detalle no estan en periodos.json; su nombre se dedujo de la fecha de vencimiento."])
        ws_resumen.cell(ws_resumen.max_row, 1).font = FONT_NOTE
        ws_resumen.append(["Nota: 'SIN AULA (no esta en el padron)' agrupa a los alumnos con deuda que ya no figuran en el padron de aulas (traslados, retiros o deuda de años anteriores)."])
        ws_resumen.cell(ws_resumen.max_row, 1).font = FONT_NOTE

        widths_resumen = {'A': 26.0, 'B': 18.0, 'C': 18.0, 'D': 18.0, 'E': 18.0, 'F': 18.0, 'G': 18.0, 'H': 18.0, 'I': 18.0}
        for col_l, w in widths_resumen.items():
            ws_resumen.column_dimensions[col_l].width = w

        # =========================================================================
        # 4. HOJAS POR SECCIÓN
        # =========================================================================
        headers_seccion = [
            "Nombre completo del alumno", "DNI estudiante", "Celular / teléfono",
            "Nombre del apoderado", "Meses que debe", "Módulos que debe", "N° de cuotas",
            "Deuda (S/)", "Mora (S/)", "Total a pagar (S/)", "Vencimiento más antiguo", "Deuda total del alumno (S/)"
        ]
        widths_seccion = {'A': 38.0, 'B': 14.0, 'C': 22.0, 'D': 34.0, 'E': 30.0, 'F': 20.0, 'G': 11.0, 'H': 13.0, 'I': 11.0, 'J': 15.0, 'K': 18.0, 'L': 18.0}

        for aula_nom, a_info in aulas_dict.items():
            sheet_title = aula_nom[:31]
            ws_sec = wb.create_sheet(sheet_title)
            ws_sec.views.sheetView[0].showGridLines = True

            ws_sec.merge_cells("A1:L1")
            ws_sec["A1"] = f"ALUMNOS DEUDORES — {aula_nom.upper()}"
            ws_sec["A1"].font = FONT_TITLE
            ws_sec["A1"].fill = FILL_BANNER
            ws_sec["A1"].alignment = ALIGN_CENTER

            ws_sec.merge_cells("A2:L2")
            ws_sec["A2"] = f"{a_info['grado']}   |   Periodos solicitados: {periodos_texto}"
            ws_sec["A2"].font = FONT_SUBTITLE
            ws_sec["A2"].alignment = ALIGN_LEFT

            ws_sec.append([])
            ws_sec.append(headers_seccion)
            for c in range(1, 13):
                cell = ws_sec.cell(4, c)
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_CELL

            s_row = 5
            for d in a_info["alumnos"]:
                ws_sec.append([
                    d["alumno"],
                    d["dni"],
                    d["telefono"],
                    d["apoderado"],
                    d["meses_str"],
                    d["modulos_str"],
                    d["num_cuotas"],
                    d["deuda"],
                    d["mora"],
                    f"=H{s_row}+I{s_row}",
                    d["vencimiento_antiguo"],
                    d["deuda_total_alumno"]
                ])
                fill_row = FILL_ZEBRA if s_row % 2 == 0 else PatternFill(fill_type=None)
                for c in range(1, 13):
                    cell = ws_sec.cell(s_row, c)
                    cell.font = FONT_DATA
                    if fill_row.fill_type:
                        cell.fill = fill_row
                    cell.border = BORDER_CELL
                    if c in (1, 4, 5, 6):
                        cell.alignment = ALIGN_LEFT
                    elif c in (2, 3, 7, 11):
                        cell.alignment = ALIGN_CENTER
                    elif c in (8, 9, 10, 12):
                        cell.alignment = ALIGN_RIGHT
                        cell.number_format = FORMAT_CURRENCY
                s_row += 1

            tot_sec = [
                f"TOTAL ({len(a_info['alumnos'])} alumnos)",
                None, None, None, None, None,
                f"=SUM(G5:G{s_row-1})",
                f"=SUM(H5:H{s_row-1})",
                f"=SUM(I5:I{s_row-1})",
                f"=SUM(J5:J{s_row-1})",
                None,
                f"=SUM(L5:L{s_row-1})"
            ]
            ws_sec.append(tot_sec)
            for c in range(1, 13):
                cell = ws_sec.cell(s_row, c)
                cell.font = FONT_TOTAL
                cell.fill = FILL_BANNER
                cell.border = BORDER_TOTAL
                cell.alignment = ALIGN_RIGHT
                if c in (8, 9, 10, 12):
                    cell.number_format = FORMAT_CURRENCY

            for col_l, w in widths_seccion.items():
                ws_sec.column_dimensions[col_l].width = w
            ws_sec.freeze_panes = "A5"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        nombre = f"Lista Deudores - {hoy.strftime('%d-%m-%Y')}_{anio}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
        )

    except (ProgrammingError, OperationalError):
        db.rollback()
        raise _sin_tablas(Exception())
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"No se pudo generar el reporte Excel: {e}")
