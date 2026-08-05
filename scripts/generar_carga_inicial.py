"""
Genera el script SQL de carga inicial del año escolar 2026.

Lee las tres fuentes en bruto del colegio y produce un único archivo .sql con
todos los INSERT en el orden correcto de dependencias:

  1. "3. Directorio 2026 GENERAL.pdf"      -> personal, cursos, tutorías, carga
  2. "2026 AV VACANTES Y MATRÍCULAS.xlsx"  -> alumnos, apoderados, matrículas
  3. "PAGOS CUENTA RECAUDADORA BCP AV.xlsx"-> pagos

El SQL generado NO borra ni modifica nada: son solo INSERT, pensados para
correr sobre una base de datos nueva y vacía con la misma estructura.

Uso:
    python scripts/generar_carga_inicial.py
    python scripts/generar_carga_inicial.py --salida otra_ruta.sql
"""
import os
import re
import sys
import argparse
import datetime as dt
from decimal import Decimal, InvalidOperation

import openpyxl
import bcrypt

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, RAIZ)
from datos_personal import (  # noqa: E402
    ADMINISTRATIVOS, PSICOLOGOS, AUXILIARES, DOCENTES,
    TUTORES, ASIGNACIONES, AREAS_CURSO,
)
# Misma función que usa la app al crear usuarios: los usernames generados aquí
# quedan idénticos a los que produciría el panel de administración.
from app.core.util.usuarios import generar_username  # noqa: E402

BASE_DATOS = r"C:\Users\jesus\OneDrive\Documentos\Archivos Proyecto Amancio Varona\documentos y datos"
XLS_ALUMNOS = "2026 AV VACANTES Y MATRÍCULAS.xlsx"
XLS_PAGOS = "PAGOS CUENTA RECAUDADORA BCP AV.xlsx"

ANIO = "2026"
# Diferencia máxima que se da por saldada al cruzar lo pagado con lo esperado.
# Evita crear cargos residuales de pocos soles por descuentos o redondeos.
TOLERANCIA_SALDO = Decimal("10.00")
# bcrypt con coste 10: lo verifica igual `verify_password` (el coste va dentro
# del propio hash) y hace viable generar ~630 usuarios en un tiempo razonable.
COSTE_BCRYPT = 10

avisos = []


# ══════════════════════════════════════════════════════════════════════════
#  Utilidades
# ══════════════════════════════════════════════════════════════════════════
def sql(valor):
    """Convierte un valor de Python a literal SQL seguro."""
    if valor is None or valor == "":
        return "NULL"
    if isinstance(valor, bool):
        return "1" if valor else "0"
    if isinstance(valor, (int, float, Decimal)):
        return str(valor)
    if isinstance(valor, dt.datetime):
        return "'" + valor.strftime("%Y-%m-%d %H:%M:%S") + "'"
    if isinstance(valor, dt.date):
        return "'" + valor.strftime("%Y-%m-%d") + "'"
    texto = str(valor).replace("\\", "\\\\").replace("'", "''")
    return "'" + texto + "'"


def limpiar(texto, maximo=None):
    if texto is None:
        return None
    t = " ".join(str(texto).replace("\n", " ").split()).strip()
    if not t or t.upper() in ("#N/A", "NONE", "-"):
        return None
    return t[:maximo] if maximo else t


def limpiar_dni(valor):
    """Normaliza un DNI a 8 dígitos. Devuelve None si no es utilizable."""
    if valor is None:
        return None
    t = str(valor).strip()
    if t.endswith(".0"):
        t = t[:-2]
    t = "".join(c for c in t if c.isdigit())
    if not t:
        return None
    return t.zfill(8)[:8]


def limpiar_telefono(valor):
    """La columna telefono es String(9): se toman los últimos 9 dígitos."""
    if valor is None:
        return None
    d = "".join(c for c in str(valor) if c.isdigit())
    if not d:
        return None
    return d[-9:]


def a_decimal(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    t = str(valor).strip().replace(",", "")
    try:
        return Decimal(t)
    except (InvalidOperation, ValueError):
        return None


def partir_nombre(completo):
    """
    Separa 'APELLIDOS Y NOMBRES' según la convención peruana: los dos primeros
    términos son apellidos y el resto nombres. No es infalible con apellidos
    compuestos ('DE LA CRUZ'), por eso se tratan las partículas.
    """
    partes = limpiar(completo).split()
    if len(partes) <= 2:
        return " ".join(partes), ""
    particulas = {"DE", "DEL", "LA", "LAS", "LOS", "DA", "DI", "VAN", "MC"}
    i = 0
    apellidos = []
    while i < len(partes) and len(apellidos) < 2:
        actual = [partes[i]]
        while partes[i].upper() in particulas and i + 1 < len(partes):
            i += 1
            actual.append(partes[i])
        apellidos.append(" ".join(actual))
        i += 1
    nombres = partes[i:]
    if not nombres:  # nombre demasiado corto: se deja todo como apellido
        return " ".join(partes), ""
    return " ".join(apellidos), " ".join(nombres)


def hash_password(clave):
    return bcrypt.hashpw(clave.encode("utf-8"), bcrypt.gensalt(COSTE_BCRYPT)).decode("utf-8")


# ══════════════════════════════════════════════════════════════════════════
#  Estructura académica
# ══════════════════════════════════════════════════════════════════════════
NIVELES = {"PRIMARIA": 1, "SECUNDARIA": 2}

GRADOS = []  # (id, id_nivel, nombre, orden)
for i, g in enumerate(["1ero", "2do", "3ero", "4to", "5to", "6to"], start=1):
    GRADOS.append((i, 1, g, i))
for i, g in enumerate(["1ero", "2do", "3ero", "4to", "5to"], start=1):
    GRADOS.append((6 + i, 2, g, i))

ID_GRADO = {(n, g): i for i, nid, g, _ in GRADOS for n in [("PRIMARIA" if nid == 1 else "SECUNDARIA")]}

# Nombre de hoja del Excel -> (nivel, grado, seccion)
HOJA_A_SECCION = {
    "1° Amarillo Primaria 2026": ("PRIMARIA", "1ero", "Amarillo"),
    "1° Azul Primaria 2026": ("PRIMARIA", "1ero", "Azul"),
    "2° Amarillo Primaria 2026": ("PRIMARIA", "2do", "Amarillo"),
    "2° Azul Primaria 2026": ("PRIMARIA", "2do", "Azul"),
    "3° Amarillo Primaria 2026": ("PRIMARIA", "3ero", "Amarillo"),
    "3° Azul Primaria 2026": ("PRIMARIA", "3ero", "Azul"),
    "4° Amarillo Primaria 2026": ("PRIMARIA", "4to", "Amarillo"),
    "4° Azul Primaria 2026": ("PRIMARIA", "4to", "Azul"),
    "5° Amarillo Primaria 2026": ("PRIMARIA", "5to", "Amarillo"),
    "5° Azul Primaria 2026": ("PRIMARIA", "5to", "Azul"),
    "6to Amarillo Primaria 2026": ("PRIMARIA", "6to", "Amarillo"),
    "6to Azul Primaria 2026": ("PRIMARIA", "6to", "Azul"),
    "1° A NUEVOS SECUNDARIA 2026": ("SECUNDARIA", "1ero", "A"),
    " 1° B  AMANCIO Secundaria 2026": ("SECUNDARIA", "1ero", "B"),
    "2° A Secundaria 2026": ("SECUNDARIA", "2do", "A"),
    "  2° B Secundaria 2026": ("SECUNDARIA", "2do", "B"),
    "3° A de secundaria 2026": ("SECUNDARIA", "3ero", "A"),
    "3° B de secundaria 2026": ("SECUNDARIA", "3ero", "B"),
    "4° A Secundaria 2026": ("SECUNDARIA", "4to", "A"),
    "4° B Secundaria 2026": ("SECUNDARIA", "4to", "B"),
    "5° A Secundaria 2026": ("SECUNDARIA", "5to", "A"),
    "5° B Secundaria 2026": ("SECUNDARIA", "5to", "B"),
}

SECCIONES = []  # (id, id_grado, nivel, grado, nombre)
_id_sec = 0
_visto = {}
for hoja, (niv, gr, sec) in HOJA_A_SECCION.items():
    clave = (niv, gr, sec)
    if clave in _visto:
        continue
    _id_sec += 1
    _visto[clave] = _id_sec
    SECCIONES.append((_id_sec, ID_GRADO[(niv, gr)], niv, gr, sec))
ID_SECCION = _visto


# ══════════════════════════════════════════════════════════════════════════
#  Lectura de alumnos
# ══════════════════════════════════════════════════════════════════════════
def leer_alumnos():
    ruta = os.path.join(BASE_DATOS, XLS_ALUMNOS)
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    alumnos = []
    dnis_vistos = {}
    sin_dni = 0

    for hoja, (niv, gr, sec) in HOJA_A_SECCION.items():
        if hoja not in wb.sheetnames:
            avisos.append(f"No se encontró la hoja {hoja!r} en el Excel de alumnos")
            continue
        ws = wb[hoja]
        for n_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            fila = list(fila) + [None] * 30
            nombre_completo = limpiar(fila[3], 400)
            if not nombre_completo:
                continue
            # Filas de totales o encabezados repetidos
            if nombre_completo.upper().startswith(("TOTAL", "APELLIDOS")):
                continue

            dni = limpiar_dni(fila[2])
            if not dni:
                sin_dni += 1
                avisos.append(f"[{hoja} fila {n_fila}] {nombre_completo}: sin DNI, se omite")
                continue
            if dni in dnis_vistos:
                avisos.append(
                    f"[{hoja} fila {n_fila}] DNI {dni} repetido "
                    f"({nombre_completo} vs {dnis_vistos[dni]}): se omite el segundo"
                )
                continue
            dnis_vistos[dni] = nombre_completo

            apellidos, nombres = partir_nombre(nombre_completo)
            enfermedad = limpiar(fila[4], 150)
            if enfermedad and enfermedad.upper() == "NINGUNA":
                enfermedad = None

            alumnos.append({
                "dni": dni,
                "apellidos": apellidos,
                "nombres": nombres,
                "enfermedad": enfermedad,
                "telefono_apoderado": limpiar_telefono(fila[5]),
                "apoderado": limpiar(fila[6], 400),
                "dni_apoderado": limpiar_dni(fila[7]),
                "direccion": limpiar(fila[8], 300),
                "colegio": limpiar(fila[9], 100),
                "talla": limpiar(fila[10], 5),
                "hermanos": (limpiar(fila[13]) or "").upper().startswith("S"),
                # Columna PENSIÓN: 240 con descuento por hermanos, 250 normal
                "pension": a_decimal(fila[14]) or Decimal("250.00"),
                "nivel": niv, "grado": gr, "seccion": sec,
            })
    wb.close()
    if sin_dni:
        avisos.append(f"TOTAL de alumnos omitidos por no tener DNI: {sin_dni}")
    return alumnos


# ══════════════════════════════════════════════════════════════════════════
#  Lectura de pagos
# ══════════════════════════════════════════════════════════════════════════
# Tipos de pago tal como los tiene configurados el colegio.
#
# IMPORTANTE: fecha_vencimiento debe ser POSTERIOR a fecha_inicio comparando
# las cadenas "MM-DD". El schema TipoPagoBase lo valida en cada respuesta, así
# que un rango que cruce el fin de año (12-02 -> 01-31) hace fallar el endpoint
# /finance/tipos-pago entero, no solo esa fila.
CATEGORIAS_TIPO_PAGO = [
    (9, "VACANTE", "Vacante", "100.00", "01-01", "10-10", "0.00", "DESHABILITAR"),
    (10, "MATRICULA", "Matricula", "200.00", "01-01", "10-10", "0.00", "DESHABILITAR"),
    (11, "PENSION", "Pensión regular", "250.00", "01-01", "01-30", "5.00", "APLICAR_MORA"),
    (12, "MODULO", "Modulo 1", "150.00", "03-09", "03-30", "5.00", "APLICAR_MORA"),
    (13, "MODULO", "Modulo 2", "150.00", "03-09", "07-30", "5.00", "APLICAR_MORA"),
]

# Identificadores usados al clasificar cada pago
ID_VACANTE, ID_MATRICULA, ID_PENSION, ID_MODULO_1, ID_MODULO_2 = 9, 10, 11, 12, 13


MESES_ES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
            "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

# Calendario de cobros del año escolar 2026, tomado de la hoja REGISTRO del
# Excel del BCP. Cada entrada: (clave, id_tipo_pago, concepto, monto, vencimiento)
# El monto de las pensiones se toma de cada alumno (240 con hermanos, 250 sin).
CALENDARIO_2026 = [
    ("MATRICULA|2026", ID_MATRICULA, "Matrícula 2026", Decimal("200.00"), dt.date(2026, 1, 31)),
    ("PENSION|2026|3", ID_PENSION, "Pensión marzo 2026", None, dt.date(2026, 3, 31)),
    ("MODULO|2026|I", ID_MODULO_1, "Módulos I 2026", Decimal("150.00"), dt.date(2026, 4, 10)),
    ("PENSION|2026|4", ID_PENSION, "Pensión abril 2026", None, dt.date(2026, 4, 30)),
    ("PENSION|2026|5", ID_PENSION, "Pensión mayo 2026", None, dt.date(2026, 5, 31)),
    ("PENSION|2026|6", ID_PENSION, "Pensión junio 2026", None, dt.date(2026, 6, 30)),
    ("MODULO|2026|II", ID_MODULO_2, "Módulos II 2026", Decimal("150.00"), dt.date(2026, 7, 10)),
    ("PENSION|2026|7", ID_PENSION, "Pensión julio 2026", None, dt.date(2026, 7, 31)),
    ("PENSION|2026|8", ID_PENSION, "Pensión agosto 2026", None, dt.date(2026, 8, 31)),
    ("PENSION|2026|9", ID_PENSION, "Pensión septiembre 2026", None, dt.date(2026, 9, 30)),
    ("PENSION|2026|10", ID_PENSION, "Pensión octubre 2026", None, dt.date(2026, 10, 31)),
    ("PENSION|2026|11", ID_PENSION, "Pensión noviembre 2026", None, dt.date(2026, 11, 30)),
    ("PENSION|2026|12", ID_PENSION, "Pensión diciembre 2026", None, dt.date(2026, 12, 25)),
]


def clave_cargo(mes):
    """
    Clave normalizada del cargo al que corresponde un pago, para poder cruzar
    lo que el alumno ya pagó contra lo que se le debía cobrar.
    """
    if isinstance(mes, dt.datetime):
        return f"PENSION|{mes.year}|{mes.month}"
    texto = limpiar(mes)
    if not texto:
        return None
    arriba = texto.upper()
    anio = "2026" if "2026" in arriba else ("2025" if "2025" in arriba else "")
    if "VACANTE" in arriba:
        return f"VACANTE|{anio}"
    if "MATRIC" in arriba or "MATRÍC" in arriba:
        # "MATRÍCULA 2026 - I" y "- II" son cuotas del mismo cargo
        return f"MATRICULA|{anio}"
    if "MODULO" in arriba or "MÓDULO" in arriba:
        # "II" debe ir como palabra suelta: buscar " 2" fallaba porque el año
        # " 2026" lo contiene, y todos los módulos I se leían como II.
        romano = "II" if re.search(r"\bII\b", arriba) else "I"
        return f"MODULO|{anio}|{romano}"
    for i, m in enumerate(MESES_ES, start=1):
        if m.upper() in arriba or (m == "septiembre" and "SETIEMBRE" in arriba):
            return f"PENSION|{anio}|{i}"
    return None


def clasificar_concepto(mes):
    """Devuelve (id_tipo_pago, concepto legible) a partir de la columna MES."""
    if isinstance(mes, dt.datetime):
        return ID_PENSION, f"Pensión {MESES_ES[mes.month - 1]} {mes.year}"
    texto = limpiar(mes)
    if not texto:
        return ID_PENSION, "Pago sin concepto registrado"
    arriba = texto.upper()
    if "VACANTE" in arriba:
        return ID_VACANTE, texto
    if "MATRIC" in arriba or "MATRÍC" in arriba:
        return ID_MATRICULA, texto
    if "MODULO" in arriba or "MÓDULO" in arriba:
        return (ID_MODULO_2 if re.search(r"II", arriba) else ID_MODULO_1), texto
    if any(m.upper() in arriba for m in MESES_ES) or "SETIEMBRE" in arriba:
        return ID_PENSION, texto
    return ID_PENSION, texto


def leer_pagos(dnis_validos):
    ruta = os.path.join(BASE_DATOS, XLS_PAGOS)
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    ws = wb["PAGOS CUENTA RECAUDADORA"]
    pagos = []
    sin_codigo = 0
    sin_alumno = 0
    sin_monto = 0

    for fila in ws.iter_rows(min_row=2, values_only=True):
        fila = list(fila) + [None] * 20
        if not limpiar(fila[2]):
            continue
        dni = limpiar_dni(fila[4])
        if not dni:
            sin_codigo += 1
            continue
        if dni not in dnis_validos:
            # Pago de un alumno que no está matriculado en 2026 (egresado,
            # retirado o de otra promoción): no hay a quién enlazarlo.
            sin_alumno += 1
            continue
        monto = a_decimal(fila[7])
        if monto is None:
            sin_monto += 1
            continue
        mora = a_decimal(fila[8]) or Decimal("0.00")
        total = a_decimal(fila[9]) or (monto + mora)
        # El schema PagoBase exige monto_total == monto + mora. El Excel del BCP
        # a veces trae un total mayor sin reflejarlo en la columna de mora (son
        # los S/5 de recargo), y esas filas hacían fallar /finance/pagos/ entero.
        # Se respeta el total realmente cobrado y se ajusta la mora.
        if total != monto + mora:
            if total >= monto:
                mora = total - monto
            else:
                monto, mora = total, Decimal("0.00")
        fecha_pago = fila[5] if isinstance(fila[5], dt.datetime) else None
        vence = fila[6].date() if isinstance(fila[6], dt.datetime) else None
        id_tipo, concepto = clasificar_concepto(fila[3])
        operacion = limpiar(fila[12], 50)

        pagos.append({
            "dni": dni, "id_tipo_pago": id_tipo, "concepto": concepto[:150],
            "monto": monto, "mora": mora, "monto_total": total,
            "codigo_operacion_bcp": operacion,
            "fecha_pago": fecha_pago, "fecha_vencimiento": vence,
            "clave": clave_cargo(fila[3]),
        })
    wb.close()
    avisos.append(f"Pagos sin código de alumno legible: {sin_codigo}")
    avisos.append(f"Pagos de alumnos que no están en la nómina 2026: {sin_alumno} (no se cargan)")
    avisos.append(f"Pagos con monto no numérico (p. ej. 'EXONERADO'): {sin_monto}")
    return pagos


def calcular_pendientes(alumnos, pagos_reales, hoy):
    """
    Cruza el calendario de cobros 2026 con lo que cada alumno ya pagó y genera
    los cargos que le faltan. Si pagó de más o completó el cargo en cuotas, no
    se genera nada; si pagó una parte, se genera el saldo.
    """
    # Cuánto lleva pagado cada alumno por cada cargo
    pagado = {}
    for p in pagos_reales:
        if not p["clave"]:
            continue
        clave = (p["dni"], p["clave"])
        pagado[clave] = pagado.get(clave, Decimal("0")) + p["monto_total"]

    pendientes = []
    for a in alumnos:
        for clave_cargo_, id_tipo, concepto, monto_fijo, vence in CALENDARIO_2026:
            esperado = monto_fijo if monto_fijo is not None else a["pension"]
            ya = pagado.get((a["dni"], clave_cargo_), Decimal("0"))
            saldo = esperado - ya
            # Tolerancia: la pensión de cada alumno se toma del Excel, pero el
            # colegio a veces cobró unos soles menos (descuentos, redondeos).
            # Sin este margen se generaban cientos de cargos de S/5 y S/10 que
            # no son deuda real, sino el desfase con el importe que se asumió.
            if saldo <= TOLERANCIA_SALDO:
                continue
            pendientes.append({
                "dni": a["dni"], "id_tipo_pago": id_tipo,
                "concepto": concepto, "monto": saldo,
                "mora": Decimal("0.00"), "monto_total": saldo,
                "fecha_vencimiento": vence,
                # Ya vencido -> VENCIDO; aún por vencer -> PENDIENTE
                "estado": "VENCIDO" if vence < hoy else "PENDIENTE",
            })
    return pendientes


# ══════════════════════════════════════════════════════════════════════════
#  Generación del SQL
# ══════════════════════════════════════════════════════════════════════════
def generar(salida):
    print("Leyendo nómina de alumnos...")
    alumnos = leer_alumnos()
    print(f"  {len(alumnos)} alumnos válidos")

    dnis_alumnos = {a["dni"] for a in alumnos}
    print("Leyendo pagos...")
    pagos = leer_pagos(dnis_alumnos)
    print(f"  {len(pagos)} pagos realizados")

    hoy = dt.date.today()
    pendientes = calcular_pendientes(alumnos, pagos, hoy)
    print(f"  {len(pendientes)} cargos pendientes generados")

    # --- Asignación de IDs -------------------------------------------------
    id_usuario = 0
    usuarios = []          # (id, username, hash, rol)
    filas_admin, filas_psico, filas_aux, filas_doc = [], [], [], []
    id_por_dni_docente = {}
    usuario_por_dni = {}

    def nuevo_usuario(dni, clave, rol):
        """Crea el usuario con el username prefijado por rol (ej. DOC-74634911)."""
        nonlocal id_usuario
        id_usuario += 1
        usuarios.append((id_usuario, generar_username(dni, rol), hash_password(clave), rol))
        return id_usuario

    print("Generando usuarios del personal (bcrypt, puede tardar)...")

    def usuario_de(dni, rol):
        """
        Un usuario por cada función que cumple la persona. Gracias al prefijo
        de rol en el username, quien es docente y auxiliar tiene dos cuentas
        (DOC-xxxx y AUX-xxxx) y entra al panel que corresponda.
        """
        if not dni:
            return None
        clave = (dni, rol)
        if clave not in usuario_por_dni:
            usuario_por_dni[clave] = nuevo_usuario(dni, dni, rol)
        return usuario_por_dni[clave]

    # Administradores
    id_admin = 0
    for dni, ape, nom, tel, cargo in ADMINISTRATIVOS:
        if not dni:
            avisos.append(f"Administrativo sin DNI, se omite: {ape} {nom} ({cargo})")
            continue
        id_admin += 1
        filas_admin.append((id_admin, usuario_de(dni, "ADMIN"), dni, nom, ape, tel, cargo))

    # Psicólogos
    id_psico = 0
    for dni, ape, nom, tel in PSICOLOGOS:
        id_psico += 1
        filas_psico.append((id_psico, usuario_de(dni, "PSICOLOGO"), dni, nom, ape, tel))

    # Auxiliares
    id_aux = 0
    for dni, ape, nom, tel, ambito, turno in AUXILIARES:
        id_aux += 1
        filas_aux.append((id_aux, usuario_de(dni, "AUXILIAR"), dni, nom, ape, tel, ambito, turno))

    # Docentes
    id_doc = 0
    for dni, ape, nom, tel in DOCENTES:
        if not dni:
            avisos.append(f"Docente sin DNI en el directorio, se omite: {ape} {nom}")
            continue
        if dni in id_por_dni_docente:
            continue
        id_doc += 1
        id_por_dni_docente[dni] = id_doc
        filas_doc.append((id_doc, usuario_de(dni, "DOCENTE"), dni, nom, ape, tel))

    # Cursos derivados de las asignaciones
    cursos = {}
    areas = {}
    for _, niv, curso, _ in ASIGNACIONES:
        if curso not in cursos:
            cursos[curso] = len(cursos) + 1
        a = AREAS_CURSO.get(curso, "General")
        if a not in areas:
            areas[a] = len(areas) + 1

    # Plan de estudio + carga académica
    plan, carga = set(), []
    for dni, niv, curso, grados in ASIGNACIONES:
        for g in grados:
            clave_g = (niv, g)
            if clave_g not in ID_GRADO:
                avisos.append(f"Grado inexistente en la asignación: {niv} {g} ({curso})")
                continue
            plan.add((cursos[curso], ID_GRADO[clave_g]))
            if not dni or dni not in id_por_dni_docente:
                if dni:
                    avisos.append(f"Asignación con docente desconocido (DNI {dni}): {curso} {niv} {g}")
                continue
            for (n2, g2, s2), id_sec in ID_SECCION.items():
                if n2 == niv and g2 == g:
                    carga.append((id_sec, cursos[curso], id_por_dni_docente[dni]))

    # Alumnos, apoderados y matrículas
    print("Generando usuarios de alumnos (bcrypt, puede tardar)...")
    filas_alumno, filas_familiar, filas_relacion, filas_matricula = [], [], [], []
    familiar_por_dni, id_alumno_por_dni = {}, {}
    id_al = id_fam = id_rel = id_mat = 0

    for a in alumnos:
        id_al += 1
        id_alumno_por_dni[a["dni"]] = id_al
        uid = nuevo_usuario(a["dni"], a["dni"], "ALUMNO")
        filas_alumno.append((id_al, uid, a))

        if a["apoderado"]:
            dni_fam = a["dni_apoderado"]
            clave = dni_fam or f"SIN-DNI-{id_al}"
            if clave not in familiar_por_dni:
                id_fam += 1
                familiar_por_dni[clave] = id_fam
                ape_f, nom_f = partir_nombre(a["apoderado"])
                filas_familiar.append((id_fam, dni_fam, nom_f, ape_f,
                                       a["telefono_apoderado"], a["direccion"]))
            id_rel += 1
            filas_relacion.append((id_rel, id_al, familiar_por_dni[clave], "APODERADO"))

        id_mat += 1
        filas_matricula.append((
            id_mat, id_al,
            ID_SECCION[(a["nivel"], a["grado"], a["seccion"])],
            ID_GRADO[(a["nivel"], a["grado"])],
        ))

    # ═════════════════════════════════════════════════════════════════════
    print(f"Escribiendo {salida}...")
    L = []
    w = L.append

    w("-- ═══════════════════════════════════════════════════════════════════")
    w("--  CARGA INICIAL — COLEGIO AMANCIO VARONA — AÑO ESCOLAR 2026")
    w(f"--  Generado el {dt.datetime.now():%Y-%m-%d %H:%M} por scripts/generar_carga_inicial.py")
    w("--")
    w("--  Solo contiene INSERT: no borra ni modifica ningún dato existente.")
    w("--  Pensado para una base de datos NUEVA y VACÍA con la misma estructura.")
    w("--")
    w(f"--  Alumnos: {len(filas_alumno)}   Apoderados: {len(filas_familiar)}   "
      f"Matrículas: {len(filas_matricula)}")
    w(f"--  Docentes: {len(filas_doc)}   Administrativos: {len(filas_admin)}   "
      f"Auxiliares: {len(filas_aux)}   Psicólogos: {len(filas_psico)}")
    w(f"--  Cursos: {len(cursos)}   Carga académica: {len(carga)}")
    w(f"--  Pagos realizados: {len(pagos)}   Cargos por cobrar: {len(pendientes)}")
    w("--")
    w("--  NO incluye el contenido de la página pública (textos, imágenes,")
    w("--  noticias y calendario): eso no sale del PDF ni de los Excel, sino")
    w("--  del panel. Se migra aparte con scripts/migrar_contenido_web.py")
    w("--")
    w("--  Contraseña inicial de todos los usuarios = su propio DNI, y todos")
    w("--  quedan marcados para cambiarla en el primer ingreso.")
    w("--  El username lleva el prefijo del rol: ADM- DOC- ALU- AUX- PSI-")
    w("--  Ej: el director Tomás Serquén entra como ADM-16793446 y, como")
    w("--  docente, con DOC-16793446. Son dos cuentas distintas.")
    w("-- ═══════════════════════════════════════════════════════════════════")
    w("")
    w("SET NAMES utf8mb4;")
    w("SET SESSION sql_mode = 'STRICT_ALL_TABLES';")
    w("START TRANSACTION;")
    w("")

    def bloque(titulo):
        w("")
        w("-- " + "─" * 68)
        w(f"--  {titulo}")
        w("-- " + "─" * 68)

    # Un INSERT con miles de filas supera max_allowed_packet (el de pagos
    # rondaba los 16 MB y el servidor cerraba la conexión). Se parte en lotes.
    TAM_LOTE = 400

    def insertar(cabecera, filas, tam=TAM_LOTE):
        """Escribe uno o varios INSERT de `cabecera` con las filas indicadas."""
        filas = list(filas)
        if not filas:
            w(f"-- (sin filas para {cabecera.split('(')[0].strip()})")
            return
        for inicio in range(0, len(filas), tam):
            trozo = filas[inicio:inicio + tam]
            if len(filas) > tam:
                w(f"-- filas {inicio + 1} a {inicio + len(trozo)} de {len(filas)}")
            w(cabecera + " VALUES")
            w(",\n".join(trozo) + ";")
            w("")

    # 1. Año escolar
    bloque("1. AÑO ESCOLAR")
    w("INSERT INTO anio_escolar (id_anio_escolar, fecha_inicio, fecha_fin, activo, tipo, "
      "inicio_inscripcion, fin_inscripcion) VALUES")
    w(f"  ({sql(ANIO)}, '2026-03-01', '2026-12-20', 1, 'REGULAR', '2025-12-01', '2026-03-31');")

    # 2. Niveles y grados
    bloque("2. NIVELES Y GRADOS")
    w("INSERT INTO nivel (id_nivel, nombre) VALUES")
    w("  (1, 'PRIMARIA'),")
    w("  (2, 'SECUNDARIA');")
    w("")
    w("INSERT INTO grado (id_grado, id_nivel, nombre, orden) VALUES")
    w(",\n".join(f"  ({i}, {n}, {sql(g)}, {o})" for i, n, g, o in GRADOS) + ";")

    # 3. Secciones
    bloque("3. SECCIONES DEL AÑO 2026")
    w("-- Nota: seccion.nombre es VARCHAR(5); 'Amarillo' y 'Azul' se abrevian AMAR/AZUL.")
    w("INSERT INTO seccion (id_seccion, id_grado, id_anio_escolar, nombre, vacantes) VALUES")
    w(",\n".join(
        f"  ({i}, {gid}, {sql(ANIO)}, {sql(nom)}, 40)" for i, gid, _, _, nom in SECCIONES
    ) + ";")

    # 4. Áreas y cursos
    bloque("4. ÁREAS, CURSOS Y PLAN DE ESTUDIO")
    w("INSERT INTO area (id_area, nombre) VALUES")
    w(",\n".join(f"  ({i}, {sql(n)})" for n, i in sorted(areas.items(), key=lambda x: x[1])) + ";")
    w("")
    w("INSERT INTO curso (id_curso, id_area, nombre, minutos_semanales, es_verano) VALUES")
    w(",\n".join(
        f"  ({i}, {areas[AREAS_CURSO.get(n, 'General')]}, {sql(n)}, 0, 0)"
        for n, i in sorted(cursos.items(), key=lambda x: x[1])
    ) + ";")
    w("")
    insertar(
        "INSERT INTO plan_estudio (id_plan_estudio, id_curso, id_grado)",
        (f"  ({i}, {c}, {g})" for i, (c, g) in enumerate(sorted(plan), start=1)),
    )

    # 5. Tipos de pago
    bloque("5. TIPOS DE PAGO")
    w("INSERT INTO tipo_pago (id_tipo_pago, categoria, nombre, costo, fecha_inicio, "
      "fecha_vencimiento, mora, accion_vencimiento, activo, periodo_academico) VALUES")
    w(",\n".join(
        f"  ({i}, {sql(cat)}, {sql(nom)}, {costo}, {sql(ini)}, {sql(fin)}, "
        f"{mora}, {sql(accion)}, 1, 'REGULAR')"
        for i, cat, nom, costo, ini, fin, mora, accion in CATEGORIAS_TIPO_PAGO
    ) + ";")

    # 6. Usuarios
    bloque(f"6. USUARIOS ({len(usuarios)})")
    w("-- La contraseña inicial de cada usuario es su propio DNI (hash bcrypt).")
    insertar(
        "INSERT INTO usuario (id_usuario, username, password_hash, rol, activo, "
        "debe_cambiar_password)",
        (f"  ({i}, {sql(u)}, {sql(h)}, {sql(r)}, 1, 1)" for i, u, h, r in usuarios),
    )

    # 7. Personal
    bloque("7. PERSONAL")
    w("INSERT INTO administrador (id_admin, id_usuario, dni, nombres, apellidos, telefono, "
      "email, permisos) VALUES")
    # Comentario de bloque, no "--": un comentario de línea se comería la coma
    # separadora que el join añade justo después.
    w(",\n".join(
        f"  ({i}, {u}, {sql(d)}, {sql(n)}, {sql(a)}, {sql(t)}, NULL, NULL) "
        f"/* {cargo} */"
        for i, u, d, n, a, t, cargo in filas_admin
    ) + ";")
    w("")
    w("INSERT INTO psicologo (id_psicologo, id_usuario, dni, nombres, apellidos, telefono, "
      "email) VALUES")
    w(",\n".join(
        f"  ({i}, {u}, {sql(d)}, {sql(n)}, {sql(a)}, {sql(t)}, NULL)"
        for i, u, d, n, a, t in filas_psico
    ) + ";")
    w("")
    w("INSERT INTO auxiliar (id_auxiliar, id_usuario, dni, nombres, apellidos, telefono, "
      "email) VALUES")
    w(",\n".join(
        f"  ({i}, {u}, {sql(d)}, {sql(n)}, {sql(a)}, {sql(t)}, NULL) /* {amb} - {tur} */"
        for i, u, d, n, a, t, amb, tur in filas_aux
    ) + ";")
    w("")
    w("INSERT INTO docente (id_docente, id_usuario, dni, nombres, apellidos, especialidad, "
      "descripcion, telefono, email, url_perfil, visible_web) VALUES")
    w(",\n".join(
        f"  ({i}, {u}, {sql(d)}, {sql(n)}, {sql(a)}, NULL, NULL, {sql(t)}, NULL, NULL, 1)"
        for i, u, d, n, a, t in filas_doc
    ) + ";")

    # 8. Alumnos
    bloque(f"8. ALUMNOS ({len(filas_alumno)})")
    insertar(
        "INSERT INTO alumno (id_alumno, id_usuario, dni, nombres, apellidos, direccion, "
        "enfermedad, talla_polo, colegio_procedencia, id_grado_ingreso, relacion_fraternal, "
        "estado_ingreso)",
        (
            f"  ({i}, {u}, {sql(a['dni'])}, {sql(a['nombres'])}, {sql(a['apellidos'])}, "
            f"{sql(a['direccion'])}, {sql(a['enfermedad'])}, {sql(a['talla'])}, "
            f"{sql(a['colegio'])}, {ID_GRADO[(a['nivel'], a['grado'])]}, "
            f"{sql(a['hermanos'])}, 'ACEPTADO')"
            for i, u, a in filas_alumno
        ),
    )

    # 9. Apoderados
    bloque(f"9. APODERADOS ({len(filas_familiar)})")
    insertar(
        "INSERT INTO familiar (id_familiar, dni, nombres, apellidos, telefono, email, direccion)",
        (
            f"  ({i}, {sql(d)}, {sql(n)}, {sql(a)}, {sql(t)}, NULL, {sql(dir_)})"
            for i, d, n, a, t, dir_ in filas_familiar
        ),
    )
    w("-- IMPORTANTE: la columna familiar.email queda NULL porque el Excel no trae correos.")
    w("-- Sin ese dato NO se enviarán las notificaciones de asistencia a los apoderados.")
    w("")
    insertar(
        "INSERT INTO relacion_familiar (id_relacion_familiar, id_alumno, id_familiar, "
        "tipo_parentesco)",
        (f"  ({i}, {al}, {fa}, {sql(p)})" for i, al, fa, p in filas_relacion),
    )

    # 10. Matrículas
    bloque(f"10. MATRÍCULAS 2026 ({len(filas_matricula)})")
    insertar(
        "INSERT INTO matricula (id_matricula, id_anio_escolar, id_alumno, id_seccion, "
        "id_grado, fecha_matricula, estado, tipo_matricula, condicion)",
        (
            f"  ({i}, {sql(ANIO)}, {al}, {sec}, {gr}, '2026-03-01 00:00:00', "
            f"'MATRICULADO', 'REGULAR', 'NORMAL')"
            for i, al, sec, gr in filas_matricula
        ),
    )

    # 11. Tutores
    bloque("11. TUTORES DE SECCIÓN")
    filas_tutor = []
    for dni, niv, gr, sec in TUTORES:
        if dni not in id_por_dni_docente:
            avisos.append(f"Tutor con DNI {dni} no está en la lista de docentes")
            continue
        filas_tutor.append((len(filas_tutor) + 1, ID_SECCION[(niv, gr, sec)],
                            id_por_dni_docente[dni]))
    if filas_tutor:
        w("INSERT INTO tutor_seccion (id_tutor_seccion, id_anio_escolar, id_seccion, "
          "id_docente) VALUES")
        w(",\n".join(
            f"  ({i}, {sql(ANIO)}, {s}, {d})" for i, s, d in filas_tutor
        ) + ";")
    w("")
    w("-- El directorio solo indica 'MAESTRA DE AULA' para 1ero a 3ero de primaria.")
    w("-- Las demás secciones quedan sin tutor asignado.")

    # 12. Carga académica
    bloque(f"12. CARGA ACADÉMICA ({len(carga)})")
    insertar(
        "INSERT INTO carga_academica (id_carga_academica, id_anio_escolar, id_seccion, "
        "id_curso, id_docente)",
        (
            f"  ({i}, {sql(ANIO)}, {s}, {c}, {d})"
            for i, (s, c, d) in enumerate(sorted(set(carga)), start=1)
        ),
    )

    # 13. Pagos
    n_vencidos = sum(1 for p in pendientes if p["estado"] == "VENCIDO")
    n_por_vencer = len(pendientes) - n_vencidos
    bloque(f"13. PAGOS ({len(pagos)} pagados + {len(pendientes)} por cobrar)")
    w("-- Los PAGADO provienen del extracto de la cuenta recaudadora del BCP.")
    w("-- Los pendientes salen de cruzar el calendario de cobros 2026 contra lo")
    w("-- que cada alumno ya pagó: si un cargo quedó a medias, se registra el saldo.")
    w(f"-- VENCIDO: {n_vencidos}   PENDIENTE: {n_por_vencer}")

    filas_pago = []
    n = 0
    for p in pagos:
        n += 1
        filas_pago.append(
            f"  ({n}, {id_alumno_por_dni[p['dni']]}, {p['id_tipo_pago']}, "
            f"{sql(p['concepto'])}, {p['monto']}, {p['mora']}, {p['monto_total']}, "
            f"{sql(p['codigo_operacion_bcp'])}, 'PAGADO', "
            f"{sql(p['fecha_vencimiento'])}, {sql(p['fecha_pago'])})"
        )
    for p in pendientes:
        n += 1
        filas_pago.append(
            f"  ({n}, {id_alumno_por_dni[p['dni']]}, {p['id_tipo_pago']}, "
            f"{sql(p['concepto'])}, {p['monto']}, {p['mora']}, {p['monto_total']}, "
            f"NULL, {sql(p['estado'])}, {sql(p['fecha_vencimiento'])}, NULL)"
        )

    insertar(
        "INSERT INTO pago (id_pago, id_alumno, id_tipo_pago, concepto, monto, mora, "
        "monto_total, codigo_operacion_bcp, estado, fecha_vencimiento, fecha_pago)",
        filas_pago,
    )

    # 14. Configuración del sistema
    bloque("14. CONFIGURACIÓN DEL SISTEMA")
    w("-- Interruptor del cambio de contraseña obligatorio en el primer ingreso.")
    w("-- Se administra desde Panel de control > Seguridad.")
    w("INSERT INTO pagina_configuracion (seccion, clave, valor, tipo) VALUES")
    w("  ('seguridad', 'forzar_cambio_password_inicial', '1', 'text');")

    w("")
    w("COMMIT;")
    w("")
    w("-- ═══════════════════════════════════════════════════════════════════")
    w("--  AVISOS DE LA GENERACIÓN (revisar antes de dar por buena la carga)")
    w("-- ═══════════════════════════════════════════════════════════════════")
    for a in avisos:
        w(f"--  * {a}")
    w("")

    with open(salida, "w", encoding="utf-8") as f:
        f.write("\n".join(L))

    print(f"\nListo: {salida}")
    print(f"  {len(L)} líneas")
    print("\nRESUMEN:")
    print(f"  Usuarios .......... {len(usuarios)}")
    print(f"  Alumnos ........... {len(filas_alumno)}")
    print(f"  Apoderados ........ {len(filas_familiar)}")
    print(f"  Matrículas ........ {len(filas_matricula)}")
    print(f"  Docentes .......... {len(filas_doc)}")
    print(f"  Administrativos ... {len(filas_admin)}")
    print(f"  Auxiliares ........ {len(filas_aux)}")
    print(f"  Psicólogos ........ {len(filas_psico)}")
    print(f"  Cursos ............ {len(cursos)}")
    print(f"  Carga académica ... {len(set(carga))}")
    print(f"  Tutores ........... {len(filas_tutor)}")
    print(f"  Pagos realizados .. {len(pagos)}")
    print(f"  Cargos por cobrar . {len(pendientes)}")
    print(f"\n  Avisos: {len(avisos)} (al final del .sql)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--salida", default=os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "carga_inicial_2026.sql"))
    args = ap.parse_args()
    generar(args.salida)
