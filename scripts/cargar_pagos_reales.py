"""
Reemplaza por completo la tabla `pago` con los dos archivos oficiales del colegio.

  * PAGOS CUENTA RECAUDADORA BCP AV (2025-2026).xlsx
    hoja "2025-2026 PAGOS CUENTA RECAUDAD" -> todo lo que YA se cobró.
    Cada fila es una operación real en la cuenta recaudadora del BCP.

  * SDR-MacroCREP-AMANCIO_VARONA - copia.xlsm
    hoja "Generar Archivo de Cobranza" -> lo que FALTA cobrar del 2026.
    Es el archivo que el colegio le entrega al banco, así que la deuda ya no
    se deduce por cruce como en la carga inicial: viene dictada por el colegio.

El concepto de los cargos pendientes se identifica por la fecha de vencimiento,
según la regla del colegio: vence el 10 de abril -> Módulos I, el 10 de julio ->
Módulos II, y los días 25, 30 o 31 de cualquier mes -> pensión de ese mes.
Esa fecha solo sirve para identificar; las plantillas de `tipo_pago` no se tocan.

Solo se cargan pagos de alumnos que existen en la base. Uso:

    python scripts/cargar_pagos_reales.py                 # simulación
    python scripts/cargar_pagos_reales.py --aplicar       # escribe de verdad
"""
import argparse
import collections
import datetime as dt
import os
import re
import sys
from decimal import Decimal, InvalidOperation

import openpyxl
import pymysql
from dotenv import load_dotenv

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATOS = r"C:\Users\jesus\OneDrive\Documentos\Archivos Proyecto Amancio Varona\documentos y datos"
XLS_BCP = os.path.join(DATOS, "PAGOS CUENTA RECAUDADORA BCP AV (2025-2026).xlsx")
HOJA_BCP = "2025-2026 PAGOS CUENTA RECAUDAD"
XLS_SDR = os.path.join(DATOS, "SDR-MacroCREP-AMANCIO_VARONA - copia.xlsm")
HOJA_SDR = "Generar Archivo de Cobranza"

# Plantillas ya existentes en la base; el script las usa, nunca las modifica.
ID_VACANTE, ID_MATRICULA, ID_PENSION, ID_MODULO_1, ID_MODULO_2 = 9, 10, 11, 12, 13

MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
         "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

avisos = []


# --------------------------------------------------------------- utilidades
def texto(valor, maximo=None):
    if valor is None:
        return None
    s = " ".join(str(valor).split()).strip()
    if not s:
        return None
    return s[:maximo] if maximo else s


def dinero(valor):
    """Convierte a Decimal con 2 decimales; None si no es un número."""
    if valor is None:
        return None
    if isinstance(valor, dt.datetime):
        return None
    try:
        return Decimal(str(valor).replace(",", "").strip()).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def dni_de(valor):
    """El colegio usa el DNI del alumno como código de depositante."""
    if valor is None:
        return None
    s = re.sub(r"\D", "", str(valor))
    return s if len(s) == 8 else None


def fecha_de(valor):
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    s = texto(valor)
    if not s:
        return None
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(s, formato).date()
        except ValueError:
            pass
    return None


def pension(anio, mes):
    return ID_PENSION, f"Pensión {MESES[mes - 1]} {anio}"


# ------------------------------------------------- clasificación de lo pagado
def clasificar_pagado(mes, vence, pagado_el):
    """
    Devuelve (id_tipo_pago, concepto) a partir de la columna MES del BCP, que
    unas veces trae una fecha y otras el nombre del cargo escrito a mano.
    """
    if isinstance(mes, dt.datetime):
        return pension(mes.year, mes.month)

    crudo = texto(mes)
    if not crudo:
        return None, None
    arriba = crudo.upper()
    if arriba.startswith("#N/A") or "NO PAGOS REGISTRADOS" in arriba:
        return None, None

    # El año no siempre está escrito; se completa con las fechas de la fila.
    m = re.search(r"(20\d{2})", arriba)
    if m:
        anio = int(m.group(1))
    elif vence:
        anio = vence.year
    elif pagado_el:
        anio = pagado_el.year
    else:
        return None, None

    # Pago adelantado de todo el año: se respeta el texto del colegio para que
    # se distinga de una matrícula normal al revisarlo en caja.
    if "AÑO ESCOLAR" in arriba or "ANO ESCOLAR" in arriba:
        return ID_MATRICULA, f"Año escolar {anio} + Matrícula + Módulos I"

    if "VACANTE" in arriba:
        return ID_VACANTE, f"Vacante {anio}"

    if "MATRIC" in arriba or "MATRÍC" in arriba:
        # "MATRÍCULA 2026 - I" y "- II" son las dos cuotas de la matrícula
        cuota = re.search(r"-\s*(I{1,2})\s*$", arriba)
        if cuota:
            return ID_MATRICULA, f"Matrícula {anio} - {cuota.group(1)}"
        return ID_MATRICULA, f"Matrícula {anio}"

    if "MODULO" in arriba or "MÓDULO" in arriba:
        # "II" tiene que ir como palabra suelta: buscar " 2" fallaba porque el
        # año " 2026" lo contiene y todos los módulos I salían como II.
        romano = "II" if re.search(r"\bII\b", arriba) else "I"
        return (ID_MODULO_2 if romano == "II" else ID_MODULO_1), f"Módulos {romano} {anio}"

    if "SETIEMBRE" in arriba:
        return pension(anio, 9)
    for i, nombre in enumerate(MESES, start=1):
        if nombre.upper() in arriba:
            return pension(anio, i)

    return None, None


# --------------------------------------------- clasificación de lo pendiente
def clasificar_pendiente(vence, monto):
    """
    Regla del colegio: el vencimiento identifica el cargo.
      * 10 de abril  -> Módulos I
      * 10 de julio  -> Módulos II
      * días 25/30/31 -> pensión de ese mes
    """
    if vence is None:
        return None, None
    if vence.day == 10:
        if vence.month == 4:
            return ID_MODULO_1, f"Módulos I {vence.year}"
        if vence.month == 7:
            return ID_MODULO_2, f"Módulos II {vence.year}"
        return None, None
    if vence.day in (25, 30, 31):
        return pension(vence.year, vence.month)
    # Único caso fuera de la regla en el archivo: un vencimiento a mitad de
    # febrero por el importe exacto de la matrícula.
    if monto is not None and monto == Decimal("200.00"):
        avisos.append(
            f"Cargo con vencimiento {vence} (fuera de la regla) tomado como "
            f"Matrícula {vence.year} por coincidir con su importe."
        )
        return ID_MATRICULA, f"Matrícula {vence.year}"
    return None, None


# ------------------------------------------------------------------ lectura
def leer_pagados(alumnos):
    wb = openpyxl.load_workbook(XLS_BCP, data_only=True, read_only=True)
    ws = wb[HOJA_BCP]
    filas = []
    desc = collections.Counter()

    for fila in ws.iter_rows(min_row=2, values_only=True):
        f = list(fila) + [None] * 25

        # Una fila del archivo trae el código repetido en la columna de fecha
        # de pago, lo que corre todo lo demás un lugar a la derecha.
        if not isinstance(f[5], dt.datetime) and isinstance(f[6], dt.datetime) \
                and isinstance(f[7], dt.datetime):
            f = f[:5] + f[6:]
            desc["filas realineadas"] += 1

        dni = dni_de(f[4])
        if not dni:
            desc["sin código de alumno"] += 1
            continue
        if dni not in alumnos:
            desc["alumno fuera de la base"] += 1
            continue

        monto = dinero(f[7])
        if monto is None:
            desc["importe no numérico (EXONERADO y similares)"] += 1
            continue

        pagado_el = fecha_de(f[5])
        vence = fecha_de(f[6])
        id_tipo, concepto = clasificar_pagado(f[3], vence, pagado_el)
        if not concepto:
            desc["concepto no reconocido"] += 1
            continue

        mora = dinero(f[8]) or Decimal("0.00")
        total = dinero(f[9])
        if total is None:
            total = monto + mora
        # El extracto cobra S/5 de recargo sin anotarlo en la columna MORA. El
        # schema PagoBase exige monto_total == monto + mora, y una sola fila
        # descuadrada tumba el endpoint /finance/pagos/ entero.
        if total != monto + mora:
            if total >= monto:
                mora = total - monto
                desc["recargo movido a la columna mora"] += 1
            else:
                monto, mora = total, Decimal("0.00")
                desc["importe corregido al total cobrado"] += 1

        filas.append({
            "dni": dni, "id_tipo_pago": id_tipo, "concepto": concepto[:150],
            "monto": monto, "mora": mora, "monto_total": total,
            "operacion": texto(f[12], 50), "vence": vence,
            "pagado_el": f[5] if isinstance(f[5], dt.datetime) else None,
            "estado": "PAGADO",
        })

    wb.close()
    return filas, desc


def leer_pendientes(alumnos, hoy):
    wb = openpyxl.load_workbook(XLS_SDR, data_only=True, read_only=True)
    ws = wb[HOJA_SDR]
    filas = []
    desc = collections.Counter()

    for fila in ws.iter_rows(min_row=11, values_only=True):
        f = list(fila) + [None] * 25
        if not f[0]:
            continue

        dni = dni_de(f[0])
        if not dni:
            desc["sin código de depositante"] += 1
            continue
        if dni not in alumnos:
            desc["alumno fuera de la base"] += 1
            continue

        monto = dinero(f[5])
        if monto is None:
            desc["importe no numérico"] += 1
            continue

        vence = fecha_de(f[4])
        id_tipo, concepto = clasificar_pendiente(vence, monto)
        if not concepto:
            desc["vencimiento no clasificable"] += 1
            continue

        # El archivo del banco separa el recargo en "Mora / Cargo Fijo".
        mora = dinero(f[6]) or Decimal("0.00")

        filas.append({
            "dni": dni, "id_tipo_pago": id_tipo, "concepto": concepto[:150],
            "monto": monto, "mora": mora, "monto_total": monto + mora,
            "operacion": None, "vence": vence, "pagado_el": None,
            # Sin fecha de vencimiento cumplida todavía sigue siendo PENDIENTE;
            # una vez pasada, es deuda vencida.
            "estado": "VENCIDO" if vence < hoy else "PENDIENTE",
        })

    wb.close()
    return filas, desc


# ------------------------------------------------------------------- base
def conectar(bd):
    load_dotenv(os.path.join(RAIZ, ".env"))
    host, _, puerto = (os.getenv("DB_HOST") or "localhost").strip().partition(":")
    return pymysql.connect(
        host=host, port=int(puerto or 3306), user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASS") or "", database=bd, charset="utf8mb4",
    )


def respaldar(cur, bd):
    """Vuelca la tabla `pago` actual a un .sql antes de vaciarla."""
    carpeta = os.path.join(RAIZ, "scripts", "respaldos")
    os.makedirs(carpeta, exist_ok=True)
    sello = dt.datetime.now().strftime("%Y%m%d_%H%M")
    ruta = os.path.join(carpeta, f"{bd}_pago_ANTES_recarga_{sello}.sql")

    cur.execute("SELECT * FROM pago")
    columnas = [d[0] for d in cur.description]
    filas = cur.fetchall()

    def val(v):
        if v is None:
            return "NULL"
        if isinstance(v, (int, float, Decimal)):
            return str(v)
        return "'" + str(v).replace("\\", "\\\\").replace("'", "\\'") + "'"

    with open(ruta, "w", encoding="utf8") as fh:
        fh.write(f"-- Respaldo de `pago` de {bd} tomado el {dt.datetime.now():%Y-%m-%d %H:%M}\n")
        fh.write(f"-- {len(filas)} filas. Para restaurar: DELETE FROM pago; y ejecutar esto.\n")
        fh.write("SET FOREIGN_KEY_CHECKS=0;\n")
        cabecera = f"INSERT INTO pago ({', '.join('`' + c + '`' for c in columnas)}) VALUES\n"
        for i in range(0, len(filas), 400):
            lote = filas[i:i + 400]
            fh.write(cabecera)
            fh.write(",\n".join("  (" + ", ".join(val(v) for v in f) + ")" for f in lote))
            fh.write(";\n")
        fh.write("SET FOREIGN_KEY_CHECKS=1;\n")
    return ruta, len(filas)


# ------------------------------------------------------------------ informe
def informar(titulo, desc):
    print(f"\n  {titulo}")
    for k, v in desc.most_common():
        print(f"     {v:>6}  {k}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--bd", default="amancio_2026")
    ap.add_argument("--aplicar", action="store_true")
    args = ap.parse_args()
    hoy = dt.date.today()

    con = conectar(args.bd)
    cur = con.cursor()
    cur.execute("SELECT dni, id_alumno, apellidos, nombres FROM alumno")
    alumnos = {str(r[0]).strip(): r[1] for r in cur.fetchall()}

    cur.execute("SELECT id_tipo_pago FROM tipo_pago")
    tipos = {r[0] for r in cur.fetchall()}
    faltan = {ID_VACANTE, ID_MATRICULA, ID_PENSION, ID_MODULO_1, ID_MODULO_2} - tipos
    if faltan:
        sys.exit(f"Faltan plantillas de tipo_pago en la base: {sorted(faltan)}")

    cur.execute("SELECT COUNT(*) FROM pago WHERE id_solicitud_tramite IS NOT NULL")
    ligados = cur.fetchone()[0]
    if ligados:
        sys.exit(f"Hay {ligados} pagos ligados a un trámite; no se puede vaciar la tabla a ciegas.")

    print("=" * 78)
    print(f"CARGA DE PAGOS REALES EN `{args.bd}`   ({hoy:%d/%m/%Y})")
    print(f"Alumnos en la base: {len(alumnos)}")
    print("=" * 78)

    pagados, d1 = leer_pagados(alumnos)
    pendientes, d2 = leer_pendientes(alumnos, hoy)
    informar(f"Extracto del BCP -> {len(pagados)} pagos cargados", d1)
    informar(f"Archivo de cobranza -> {len(pendientes)} cargos cargados", d2)

    # --- comprobaciones antes de escribir nada
    problemas = []
    for p in pagados + pendientes:
        if p["monto_total"] != p["monto"] + p["mora"]:
            problemas.append(f"descuadre en {p['dni']} {p['concepto']}")
        if p["monto"] < 0 or p["mora"] < 0:
            problemas.append(f"importe negativo en {p['dni']} {p['concepto']}")
        if p["estado"] == "PAGADO" and not p["pagado_el"]:
            problemas.append(f"pago sin fecha en {p['dni']} {p['concepto']}")
    if problemas:
        print("\n  PROBLEMAS DETECTADOS:")
        for x in problemas[:20]:
            print("     ", x)
        sys.exit(f"\n{len(problemas)} filas inconsistentes: no se escribe nada.")

    por_estado = collections.Counter(p["estado"] for p in pagados + pendientes)
    print("\n  Resumen de lo que se va a insertar")
    for estado in ("PAGADO", "VENCIDO", "PENDIENTE"):
        filas = [p for p in pagados + pendientes if p["estado"] == estado]
        print(f"     {estado:<10} {por_estado[estado]:>6} cargos   "
              f"S/ {sum(p['monto_total'] for p in filas):>14,.2f}")

    anios = collections.Counter()
    for p in pagados:
        anios[p["pagado_el"].year] += 1
    print("\n  Pagos por año de cobro:", dict(sorted(anios.items())))

    conceptos = collections.Counter(p["concepto"] for p in pendientes)
    print("\n  Cargos pendientes por concepto")
    for c, n in sorted(conceptos.items(), key=lambda x: -x[1]):
        print(f"     {n:>5}  {c}")

    con_deuda = {p["dni"] for p in pendientes}
    print(f"\n  Alumnos con deuda: {len(con_deuda)} de {len(alumnos)}")
    print(f"  Alumnos al día   : {len(alumnos) - len(con_deuda)}")

    # El archivo de cobranza es una foto de un día concreto: si el alumno pagó
    # después de generarlo, el cargo sigue apareciendo como deuda. Se cargan los
    # dos archivos tal cual y se avisa para que el colegio decida, en lugar de
    # borrar por cuenta propia un cargo que el colegio dice que se debe.
    ya_pagado = collections.defaultdict(list)
    for p in pagados:
        ya_pagado[(p["dni"], p["concepto"].split(" - ")[0])].append(p)
    choques = [(p, ya_pagado[(p["dni"], p["concepto"].split(" - ")[0])])
               for p in pendientes
               if (p["dni"], p["concepto"].split(" - ")[0]) in ya_pagado]
    if choques:
        print(f"\n  REVISAR: {len(choques)} cargos siguen como deuda pese a "
              f"figurar cobrados en el extracto del BCP")
        for p, previos in choques:
            cobros = ", ".join(f"S/{x['monto_total']} el {x['pagado_el']:%d/%m/%Y}"
                               for x in previos)
            print(f"     DNI {p['dni']}  {p['concepto']}  "
                  f"deuda S/{p['monto_total']}  |  cobrado: {cobros}")

    for a in avisos:
        print(f"\n  AVISO: {a}")

    if not args.aplicar:
        print("\nSimulación. Repite con --aplicar para escribir en la base.")
        con.close()
        return

    # --- escritura
    ruta, n_previos = respaldar(cur, args.bd)
    print(f"\n  Respaldo de los {n_previos} pagos actuales en:\n     {ruta}")

    filas = []
    for p in pagados + pendientes:
        filas.append((
            alumnos[p["dni"]], p["id_tipo_pago"], p["concepto"], p["monto"],
            p["mora"], p["monto_total"], p["operacion"], p["estado"],
            p["vence"], p["pagado_el"],
        ))

    try:
        cur.execute("DELETE FROM pago")
        cur.execute("ALTER TABLE pago AUTO_INCREMENT = 1")
        cur.executemany(
            "INSERT INTO pago (id_alumno, id_tipo_pago, concepto, monto, mora, "
            "monto_total, codigo_operacion_bcp, estado, fecha_vencimiento, fecha_pago) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            filas,
        )
        con.commit()
    except Exception:
        con.rollback()
        raise

    cur.execute("SELECT estado, COUNT(*), SUM(monto_total) FROM pago GROUP BY estado")
    print("\n  Estado final de la tabla `pago`")
    for estado, n, total in cur.fetchall():
        print(f"     {estado:<10} {n:>6} cargos   S/ {total:>14,.2f}")
    cur.execute("SELECT COUNT(*) FROM pago WHERE monto_total <> monto + mora")
    print(f"     filas descuadradas: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM pago p LEFT JOIN alumno a ON a.id_alumno = p.id_alumno "
                "WHERE a.id_alumno IS NULL")
    print(f"     pagos huérfanos   : {cur.fetchone()[0]}")
    con.close()


if __name__ == "__main__":
    main()
